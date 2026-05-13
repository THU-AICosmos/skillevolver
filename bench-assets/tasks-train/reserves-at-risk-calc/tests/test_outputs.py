"""
Tests for Reserves at Risk (RaR) calculation task — TRAIN VARIANT.

Verifies:
- Step 1: Gold price volatility calculations
- Step 2: Gold reserves risk calculations for each country
- Step 3: RaR as percentage of total reserves
- Formulas are used (not hardcoded Python calculations)
"""

import csv
import glob
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

EXCEL_FILE = Path("/root/output/rar_result.xlsx")
CSV_PATTERN = "/root/output/sheet.csv.*"
TOLERANCE = 0.5  # Allow 0.5 tolerance for floating point comparisons
TOLERANCE_PCT = 0.01  # Tighter tolerance for percentage values

_csv_data_cache = None
_answer_sheet_index = None


def find_answer_csv():
    """Locate the CSV file containing Answer sheet data."""
    global _answer_sheet_index

    csv_files = sorted(glob.glob(CSV_PATTERN))
    if not csv_files:
        return None

    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE, data_only=False)
        for idx, name in enumerate(wb.sheetnames):
            if "Answer" in name:
                _answer_sheet_index = idx
                wb.close()
                expected_file = f"/root/output/sheet.csv.{idx}"
                if Path(expected_file).exists():
                    return expected_file
                break
        wb.close()

    return csv_files[0] if csv_files else None


def load_csv_data():
    """Load and cache CSV data with evaluated formula values."""
    global _csv_data_cache

    if _csv_data_cache is not None:
        return _csv_data_cache

    csv_file = find_answer_csv()
    if csv_file is None:
        _csv_data_cache = {}
        return _csv_data_cache

    _csv_data_cache = {}
    try:
        with open(csv_file, encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, start=1):
                for col_idx, val in enumerate(row):
                    col_letter = chr(ord("A") + col_idx) if col_idx < 26 else None
                    if col_letter:
                        cell_ref = f"{col_letter}{row_idx}"
                        if val and val.strip():
                            try:
                                _csv_data_cache[cell_ref] = float(val)
                            except ValueError:
                                _csv_data_cache[cell_ref] = val
                        else:
                            _csv_data_cache[cell_ref] = None
    except Exception as e:
        print(f"Error loading CSV: {e}")
        _csv_data_cache = {}

    return _csv_data_cache


def get_workbook():
    """Load the workbook with data only (calculated values)."""
    return load_workbook(EXCEL_FILE, data_only=True)


def get_workbook_formulas():
    """Load the workbook with formulas."""
    return load_workbook(EXCEL_FILE, data_only=False)


def get_answer_sheet(wb):
    """Find the Answer sheet."""
    for sheet_name in wb.sheetnames:
        if "Answer" in sheet_name:
            return wb[sheet_name]
    return wb.active


def cell_value(ws, cell):
    """Get cell value, preferring xlsx direct values then falling back to CSV."""
    val = ws[cell].value
    if val is not None and isinstance(val, (int, float)):
        return val
    csv_val = cell_value_csv(cell)
    if csv_val is not None:
        return csv_val
    return val if val is not None else 0


def cell_value_csv(cell):
    """Get cell value from CSV only."""
    csv_data = load_csv_data()
    return csv_data.get(cell)


# Expected values (derived from synthetic gold price series 1995M7..2026M3)
EXPECTED_STEP1 = {
    "C3": ("z_score", 1.65),
    "C4": ("volatility_3m", 3.715280),
    "C5": ("volatility_3m_annualized", 12.870106),
    "C6": ("volatility_12m", 3.051552),
}

# Step 2: Country gold reserves and risk values (rows 11-13)
STEP2_COLS = ["C", "D", "E", "F", "G", "H", "I", "J", "K"]
EXPECTED_STEP2 = {
    "countries": ["Indonesia", "Thailand", "Philippines", "Malaysia", "Ghana", "Egypt", "Brazil", "Argentina", "Vietnam"],
    "gold": [3211.92, 12375.84, 10337.54, 3962.87, 1227.47, 7913.63, 6473.23, 2366.03, 1343.770933],
    "risk": [196.897485, 758.665152, 633.713053, 242.932308, 75.246506, 485.122247, 396.822682, 145.04264, 82.375999],
}

# Step 3: RaR as percentage of total reserves (rows 20-24)
# Excluded: Egypt (no 2026 Total Reserves), Argentina (no Total Reserves at all)
STEP3_COLS = ["C", "D", "E", "F", "G", "H", "I"]
EXPECTED_STEP3 = {
    "countries": ["Indonesia", "Thailand", "Philippines", "Malaysia", "Ghana", "Brazil", "Vietnam"],
    "gold": [3211.92, 12375.84, 10337.54, 3962.87, 1227.47, 6473.23, 1343.770933],
    "risk": [196.897485, 758.665152, 633.713053, 242.932308, 75.246506, 396.822682, 82.375999],
    "total_reserves": [238009.64, 267303.56, 170560.66, 250534.12, 10722.56, 801433.07, 124120.81],
    "rar_pct": [0.082727, 0.283822, 0.371547, 0.096966, 0.701759, 0.049514, 0.066368],
}


def test_step1_volatility_calculations():
    """Test Step 1: Gold price volatility calculations (z-score, 3m/12m volatility)."""
    assert EXCEL_FILE.exists(), f"Excel file not found at {EXCEL_FILE}"

    wb = get_workbook()
    ws = get_answer_sheet(wb)

    errors = []
    for cell, (name, expected) in EXPECTED_STEP1.items():
        actual = cell_value(ws, cell)
        if actual is None or not isinstance(actual, (int, float)):
            errors.append(f"{cell} ({name}): expected {expected}, got {actual}")
        elif abs(actual - expected) > TOLERANCE:
            errors.append(f"{cell} ({name}): expected {expected}, got {actual}")

    wb.close()
    assert len(errors) == 0, "Step 1 volatility calculation errors:\n" + "\n".join(errors)


def test_step2_gold_reserves_and_risk():
    """Test Step 2: Gold reserves values and volatility risk for each country."""
    wb = get_workbook()
    ws = get_answer_sheet(wb)

    errors = []
    for i, col in enumerate(STEP2_COLS):
        country = EXPECTED_STEP2["countries"][i]

        # Check gold reserves (row 12)
        gold_cell = f"{col}12"
        gold_expected = EXPECTED_STEP2["gold"][i]
        gold_actual = cell_value(ws, gold_cell)
        if gold_actual is None or not isinstance(gold_actual, (int, float)):
            errors.append(f"{country} gold ({gold_cell}): expected {gold_expected}, got {gold_actual}")
        elif abs(gold_actual - gold_expected) > TOLERANCE:
            errors.append(f"{country} gold ({gold_cell}): expected {gold_expected}, got {gold_actual}")

        # Check risk values (row 13)
        risk_cell = f"{col}13"
        risk_expected = EXPECTED_STEP2["risk"][i]
        risk_actual = cell_value(ws, risk_cell)
        if risk_actual is None or not isinstance(risk_actual, (int, float)):
            errors.append(f"{country} risk ({risk_cell}): expected {risk_expected}, got {risk_actual}")
        elif abs(risk_actual - risk_expected) > TOLERANCE:
            errors.append(f"{country} risk ({risk_cell}): expected {risk_expected}, got {risk_actual}")

    wb.close()
    assert len(errors) == 0, "Step 2 gold reserves/risk errors:\n" + "\n".join(errors)


def test_step3_rar_percentage():
    """Test Step 3: Gold reserves, total reserves, and RaR as percentage."""
    wb = get_workbook()
    ws = get_answer_sheet(wb)

    errors = []
    for i, col in enumerate(STEP3_COLS):
        country = EXPECTED_STEP3["countries"][i]

        # Check gold reserves (row 21)
        gold_cell = f"{col}21"
        gold_expected = EXPECTED_STEP3["gold"][i]
        gold_actual = cell_value(ws, gold_cell)
        if gold_actual is None or not isinstance(gold_actual, (int, float)):
            errors.append(f"{country} gold ({gold_cell}): expected {gold_expected}, got {gold_actual}")
        elif abs(gold_actual - gold_expected) > TOLERANCE:
            errors.append(f"{country} gold ({gold_cell}): expected {gold_expected}, got {gold_actual}")

        # Check total reserves (row 23)
        tr_cell = f"{col}23"
        tr_expected = EXPECTED_STEP3["total_reserves"][i]
        tr_actual = cell_value(ws, tr_cell)
        if tr_actual is None or not isinstance(tr_actual, (int, float)):
            errors.append(f"{country} total reserves ({tr_cell}): expected {tr_expected}, got {tr_actual}")
        elif abs(tr_actual - tr_expected) > TOLERANCE:
            errors.append(f"{country} total reserves ({tr_cell}): expected {tr_expected}, got {tr_actual}")

        # Check RaR percentage (row 24)
        rar_cell = f"{col}24"
        rar_expected = EXPECTED_STEP3["rar_pct"][i]
        rar_actual = cell_value(ws, rar_cell)
        if rar_actual is None or not isinstance(rar_actual, (int, float)):
            errors.append(f"{country} RaR% ({rar_cell}): expected {rar_expected}, got {rar_actual}")
        elif abs(rar_actual - rar_expected) > TOLERANCE_PCT:
            errors.append(f"{country} RaR% ({rar_cell}): expected {rar_expected}, got {rar_actual}")

    wb.close()
    assert len(errors) == 0, "Step 3 RaR percentage errors:\n" + "\n".join(errors)


def test_formulas_present():
    """Test that Excel formulas are used (not Python-calculated hardcoded values)."""
    assert EXCEL_FILE.exists(), f"Excel file not found at {EXCEL_FILE}"

    wb = get_workbook_formulas()

    # Check Gold price sheet has formulas for log returns and volatility
    gold_sheet = None
    for name in wb.sheetnames:
        if "gold" in name.lower() and "price" in name.lower():
            gold_sheet = wb[name]
            break
    assert gold_sheet is not None, f"Gold price sheet not found in {wb.sheetnames}"

    gold_formula_count = 0
    for row in range(3, min(50, gold_sheet.max_row + 1)):
        for col in ["C", "D", "E"]:
            cell = gold_sheet[f"{col}{row}"]
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                gold_formula_count += 1

    # Check Answer sheet has formulas
    ws = get_answer_sheet(wb)
    required_formula_cells = [
        ("C4", "3-month volatility should reference Gold price sheet"),
        ("C5", "annualized volatility should use formula"),
        ("C6", "12-month volatility should reference Gold price sheet"),
        ("C13", "gold exposure should use formula"),
        ("C22", "Step 3 gold exposure should use formula"),
        ("C24", "RaR percentage should use formula"),
    ]

    missing_formulas = []
    for cell, description in required_formula_cells:
        value = ws[cell].value
        if value is None or not (isinstance(value, str) and value.startswith("=")):
            missing_formulas.append(f"{cell}: {description} (got: {value})")

    wb.close()

    errors = []
    if gold_formula_count < 10:
        errors.append(
            f"Gold price sheet should have formulas for log returns and volatility. "
            f"Found only {gold_formula_count} formulas in columns C-E."
        )
    if missing_formulas:
        errors.append("Answer sheet missing required formulas:\n  " + "\n  ".join(missing_formulas))

    assert len(errors) == 0, (
        "Excel formulas must be used (not Python-calculated hardcoded values):\n" + "\n".join(errors)
    )


def test_no_errors_or_macros():
    """Test that there are no Excel formula errors or VBA macros."""
    assert EXCEL_FILE.exists(), f"Excel file not found at {EXCEL_FILE}"

    errors = []

    # Check for VBA macros
    with zipfile.ZipFile(EXCEL_FILE, "r") as zf:
        vba_files = [n for n in zf.namelist() if "vbaProject" in n or n.endswith(".bin")]
    if vba_files:
        errors.append(f"VBA macros not allowed: {vba_files}")

    # Check for formula errors in CSV data
    csv_data = load_csv_data()
    excel_errors = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
    error_cells = []
    for cell, val in csv_data.items():
        if val is not None and isinstance(val, str):
            for err in excel_errors:
                if err in str(val):
                    error_cells.append(f"{cell}: {val}")
                    break
    if error_cells:
        errors.append("Excel formula errors found:\n  " + "\n  ".join(error_cells[:10]))

    assert len(errors) == 0, "File validation errors:\n" + "\n".join(errors)
