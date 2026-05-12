#!/usr/bin/env python3
"""
Pytest test suite for Metabolite Abundance Analysis task.
"""

import os

import pytest
from openpyxl import load_workbook

EXCEL_FILE = "/root/metabolite_abundance.xlsx"


@pytest.fixture
def wb():
    """Load Excel workbook."""
    assert os.path.exists(EXCEL_FILE), f"Excel file not found: {EXCEL_FILE}"
    return load_workbook(EXCEL_FILE, data_only=True)


def test_file_exists():
    """Test that output Excel file exists."""
    assert os.path.exists(EXCEL_FILE)


def test_sheets_exist(wb):
    """Test that required sheets exist."""
    assert "Task" in wb.sheetnames
    assert "Data" in wb.sheetnames


def test_step1_abundance_values_filled(wb):
    """Test that abundance values are filled in Task sheet."""
    task_ws = wb["Task"]

    # Check that cells C8:N15 (8x12 region) are filled
    filled_count = 0
    for row in range(8, 16):
        for col in range(3, 15):
            value = task_ws.cell(row=row, column=col).value
            if value is not None:
                filled_count += 1

    # At least 80% of cells should be filled
    assert filled_count >= 77, f"Only {filled_count}/96 cells filled"


def test_step2_statistics_calculated(wb):
    """Test that statistics are calculated."""
    task_ws = wb["Task"]

    # Check statistics rows (19-22) have values, columns B-I
    filled_count = 0
    for row in range(19, 23):
        for col in range(2, 10):
            value = task_ws.cell(row=row, column=col).value
            if value is not None and isinstance(value, (int, float)):
                filled_count += 1

    assert filled_count >= 26, "Statistics not properly calculated"


def test_step3_fold_changes_calculated(wb):
    """Test that fold changes are calculated."""
    task_ws = wb["Task"]

    # Check Log2FC and FC columns (C and D, rows 27-34)
    log2fc_count = 0
    fc_count = 0

    for row in range(27, 35):
        log2fc_value = task_ws.cell(row=row, column=3).value
        fc_value = task_ws.cell(row=row, column=4).value

        if log2fc_value is not None and isinstance(log2fc_value, (int, float)):
            log2fc_count += 1
        if fc_value is not None and isinstance(fc_value, (int, float)):
            fc_count += 1

    assert log2fc_count >= 6, "Log2 fold changes not calculated"
    assert fc_count >= 6, "Fold changes not calculated"


def test_values_are_numeric(wb):
    """Test that filled values are numeric (not strings)."""
    task_ws = wb["Task"]

    # Check some cells in the abundance lookup area
    for row in range(8, 13):
        value = task_ws.cell(row=row, column=3).value
        if value is not None:
            assert isinstance(value, (int, float)), f"Cell ({row},3) is not numeric: {type(value)}"
