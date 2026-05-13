"""Verifier for the jpg-ocr-stat training variant (date + total).

Checks that /app/workspace/receipts_summary.xlsx contains exactly one sheet
named ``summary`` with columns (filename, receipt_date, receipt_total),
rows ordered by filename, and values matching ``receipts_oracle.xlsx``
packaged alongside this file. Implementation style intentionally differs
from the val task's verifier (uses pathlib + namedtuples + dataclasses +
pytest fixtures) so the agent can't copy-paste the val test structure.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import NamedTuple

import pytest
from openpyxl import load_workbook


OUTPUT_PATH = Path("/app/workspace/receipts_summary.xlsx")
ORACLE_PATH = Path(__file__).parent / "receipts_oracle.xlsx"
EXPECTED_SHEET = "summary"
EXPECTED_HEADER = ("filename", "receipt_date", "receipt_total")


# ---------- Row + workbook representation ------------------------------------

class Row(NamedTuple):
    filename: str
    receipt_date: str
    receipt_total: str


@dataclass(frozen=True)
class ParsedWorkbook:
    sheets: tuple[str, ...]
    header: tuple[str, ...]
    rows: tuple[Row, ...]
    raw_grid: tuple[tuple[str, ...], ...]   # full normalized 2D view (incl. header)


def _stringify(cell) -> str:
    """None -> '', everything else -> str(value)."""
    return "" if cell is None else str(cell)


def _parse(path: Path) -> ParsedWorkbook:
    wb = load_workbook(path, data_only=True)
    try:
        sheet_names = tuple(wb.sheetnames)
        first = wb[sheet_names[0]]

        n_rows = first.max_row or 0
        n_cols = first.max_column or 0

        grid: list[tuple[str, ...]] = []
        for r in range(1, n_rows + 1):
            grid.append(tuple(_stringify(first.cell(row=r, column=c).value)
                              for c in range(1, n_cols + 1)))

        if not grid:
            return ParsedWorkbook(sheets=sheet_names, header=(), rows=(), raw_grid=())

        header_row = grid[0]
        body: list[Row] = []
        for cells in grid[1:]:
            if len(cells) >= 1 and cells[0].strip() == "":
                continue   # skip blank trailing rows
            fname = cells[0] if len(cells) >= 1 else ""
            rdate = cells[1] if len(cells) >= 2 else ""
            rtotal = cells[2] if len(cells) >= 3 else ""
            body.append(Row(filename=fname, receipt_date=rdate, receipt_total=rtotal))

        return ParsedWorkbook(
            sheets=sheet_names,
            header=tuple(header_row),
            rows=tuple(body),
            raw_grid=tuple(grid),
        )
    finally:
        wb.close()


# ---------- Fixtures ---------------------------------------------------------

@pytest.fixture(scope="module")
def actual() -> ParsedWorkbook:
    assert OUTPUT_PATH.exists(), (
        f"Expected output workbook missing: {OUTPUT_PATH}"
    )
    return _parse(OUTPUT_PATH)


@pytest.fixture(scope="module")
def oracle() -> ParsedWorkbook:
    assert ORACLE_PATH.exists(), f"Oracle missing on disk: {ORACLE_PATH}"
    return _parse(ORACLE_PATH)


# ---------- Individual requirement checks ------------------------------------

def test_single_sheet_named_summary(actual: ParsedWorkbook) -> None:
    assert actual.sheets == (EXPECTED_SHEET,), (
        f"Workbook must have exactly one sheet named '{EXPECTED_SHEET}'. "
        f"Got sheets: {actual.sheets}"
    )


def test_header_matches_schema(actual: ParsedWorkbook) -> None:
    assert actual.header == EXPECTED_HEADER, (
        f"Header must be {EXPECTED_HEADER}. Got: {actual.header}"
    )


def test_no_extra_columns(actual: ParsedWorkbook) -> None:
    # Header drives column count; any wider grid means extras.
    width = max((len(r) for r in actual.raw_grid), default=0)
    assert width == len(EXPECTED_HEADER), (
        f"Workbook must have exactly {len(EXPECTED_HEADER)} columns; "
        f"detected width = {width}."
    )


def test_rows_sorted_by_filename(actual: ParsedWorkbook) -> None:
    names = [r.filename for r in actual.rows]
    assert names == sorted(names), (
        f"Data rows must be sorted by filename.\n"
        f"  got:    {names}\n"
        f"  sorted: {sorted(names)}"
    )


def test_row_count_matches_oracle(actual: ParsedWorkbook, oracle: ParsedWorkbook) -> None:
    assert len(actual.rows) == len(oracle.rows), (
        f"Expected {len(oracle.rows)} data rows; got {len(actual.rows)}."
    )


def test_filenames_match_oracle(actual: ParsedWorkbook, oracle: ParsedWorkbook) -> None:
    actual_names = [r.filename for r in actual.rows]
    expected_names = [r.filename for r in oracle.rows]
    assert actual_names == expected_names, (
        f"Filename column must match oracle exactly.\n"
        f"  actual:   {actual_names}\n"
        f"  expected: {expected_names}"
    )


def test_null_cells_align_with_oracle(actual: ParsedWorkbook, oracle: ParsedWorkbook) -> None:
    """A blank cell in the oracle must be blank in the output, and vice versa."""
    for idx, (a, e) in enumerate(zip(actual.rows, oracle.rows)):
        for field in ("receipt_date", "receipt_total"):
            a_val = getattr(a, field)
            e_val = getattr(e, field)
            a_blank = a_val.strip() == ""
            e_blank = e_val.strip() == ""
            assert a_blank == e_blank, (
                f"Null/blank pattern mismatch at row {idx + 1} "
                f"({a.filename}) field {field!r}:\n"
                f"  actual:   {a_val!r}\n"
                f"  expected: {e_val!r}"
            )


def test_receipt_dates_match_oracle(actual: ParsedWorkbook, oracle: ParsedWorkbook) -> None:
    """Every receipt_date cell must equal the oracle string verbatim
    (ISO YYYY-MM-DD)."""
    diffs: list[str] = []
    for idx, (a, e) in enumerate(zip(actual.rows, oracle.rows)):
        if a.receipt_date != e.receipt_date:
            diffs.append(
                f"  row {idx + 1} ({a.filename}): "
                f"got {a.receipt_date!r}, expected {e.receipt_date!r}"
            )
    assert not diffs, "receipt_date mismatch vs oracle:\n" + "\n".join(diffs)


def test_receipt_totals_match_oracle(actual: ParsedWorkbook, oracle: ParsedWorkbook) -> None:
    """Every receipt_total cell must equal the oracle string verbatim
    (e.g. exactly two decimals, no commas)."""
    diffs: list[str] = []
    for idx, (a, e) in enumerate(zip(actual.rows, oracle.rows)):
        if a.receipt_total != e.receipt_total:
            diffs.append(
                f"  row {idx + 1} ({a.filename}): "
                f"got {a.receipt_total!r}, expected {e.receipt_total!r}"
            )
    assert not diffs, "receipt_total mismatch vs oracle:\n" + "\n".join(diffs)
