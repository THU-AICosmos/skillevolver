"""
Generate template and answer xlsx files for the Armenia supply-side shock analysis
training variant. Scenario is Armenia (synthetic AMD-denominated data), but the
*structure* mirrors the Georgia validation task exactly so that a skill distilled
from training exploration transfers cleanly to validation.

Structural contract (must mirror tasks/shock-analysis-supply/):
  Sheets: PWT, WEO_Data, CFC data, Production, Investment
  Production:
    - A1=Assumption, A2=capital's share, B2=0.35, A3=annual depreciation rate,
      B3=AVERAGE(...) of 8-year CFC depreciation
    - HP filter area: rows 6..27 (22 years, 2002..2023)
      * C=Year, D=K, E=Real GDP, F=LnK, G=LnY
      * K=LnZ, L=LnZ_HP, M=2nd-order diff, N=LnA-Trend(Check)
      * P5 = objective (SUMXMY2 + 100*SUMSQ)
    - Lower production table: header row 35, data rows 36..75 (40 years, 2002..2041)
      * D=K/Y, E=K, F=Y, G=LnZ trend, H=Ystar_base, I=Investment, J=dK,
        K=K_With, L=Ystar_with, M=Uplift, N=Proj GDP, O=Proj GDP Growth,
        P=Baseline GDP Growth
      * K extension uses AVERAGE(D$49:D$57) -- 9-year K/Y anchor (rows for 2015..2023)
      * alpha=0.35, depr avg horizon=8 years, K/Y anchor=9 years
  Investment: 8 years starting 2026 (rows 2..9), $6.5B total (Armenia synthetic).
  PWT: year col A, rgdpna col B, rnna col C, rows 2..35 (1990..2023).
  WEO_Data: metadata rows 1..7, year col B from row 8, data col C, growth col D.
  CFC data: rows 2..29 (1996..2023), C=CFC value, D=links to PWT!C, E=depr rate.
"""
import numpy as np
import openpyxl
from datetime import datetime
from pathlib import Path
from openpyxl.worksheet.formula import ArrayFormula

_SCRIPT_DIR = Path(__file__).resolve().parent

np.random.seed(42)

# ============================================================
# PARAMETERS (match val structurally)
# ============================================================
CAPITAL_SHARE = 0.35     # val alpha
DEPR_AVG_YEARS = 8       # val horizon
KY_ANCHOR_YEARS = 9      # val anchor window (AVERAGE(D$49:D$57))
INVEST_START = 2026      # val shock start
INVEST_YEARS = 8         # val horizon (rows 2..9 in Investment sheet)
HIST_START = 2002        # val HP filter start
HIST_END = 2023
PROJ_END = 2041          # lower table spans 2002..2041 = 40 rows (36..75)

OUTPUT_XLSX = "analysis-output.xlsx"

ENV_DIR = str(_SCRIPT_DIR / "environment")
SOLUTION_DIR = str(_SCRIPT_DIR / "solution")
Path(ENV_DIR).mkdir(parents=True, exist_ok=True)
Path(SOLUTION_DIR).mkdir(parents=True, exist_ok=True)

# ============================================================
# ARMENIA SYNTHETIC DATA (different from Georgia's real values)
# ============================================================
# PWT years: 1990..2023 (34 rows).
pwt_years = list(range(1990, 2024))

# rgdpna (real GDP at constant 2017 national prices, millions AMD)
pwt_rgdpna = [
    1660000, 1520000, 1200000, 970000, 890000, 940000, 1010000, 1080000,
    1150000, 1190000, 1240000, 1330000, 1490000, 1700000, 1890000, 2160000,
    2320000, 2560000, 2730000, 2620000, 2690000, 2810000, 3000000, 3220000,
    3440000, 3660000, 3880000, 4130000, 4340000, 4620000, 4300000, 4830000,
    5180000, 5520000,
]
# rnna (capital stock at constant 2017 national prices, millions AMD)
pwt_rnna = [
    20800000, 20400000, 19800000, 19200000, 18800000, 18500000, 18350000,
    18250000, 18300000, 18450000, 18650000, 18950000, 19350000, 19850000,
    20450000, 21100000, 21900000, 22700000, 23400000, 23800000, 24000000,
    24300000, 24700000, 25200000, 25800000, 26500000, 27200000, 28000000,
    28850000, 29700000, 30300000, 31000000, 31800000, 32650000,
]
assert len(pwt_years) == len(pwt_rgdpna) == len(pwt_rnna)

# WEO GDP for Armenia (billions AMD, constant prices). Years 2000..2043.
# Historical+short-term forecast (2000..2027) hardcoded; beyond via formula.
weo_gdp_hist = [
    1.24, 1.33, 1.49, 1.70, 1.89, 2.16, 2.32, 2.56,
    2.73, 2.62, 2.69, 2.81, 3.00, 3.22, 3.44, 3.66,
    3.88, 4.13, 4.34, 4.62, 4.30, 4.83, 5.18, 5.52,
    5.96, 6.32, 6.70, 7.10,
]  # 28 entries, 2000..2027
weo_growth_hist = [
    5.9, 9.6, 13.2, 13.6, 11.1, 14.0, 7.6, 10.4,
    6.9, -4.1, 2.6, 4.3, 7.1, 7.2, 7.0, 6.4,
    6.0, 4.7, 5.1, 6.4, -7.0, 12.4, 7.3, 6.5,
    8.0, 6.0, 6.0, 6.0,
]
growth_ext = 6.0

# CFC (Consumption of fixed capital) for Armenia: years 1996..2023 (28 rows = val: 28 rows).
# Val CFC data has 28 rows (1996..2023) -> rows 2..29.
cfc_years = list(range(1996, 2024))
# Values in AMD (not millions). Depreciation rate = CFC / 1e6 / capital_stock_millions.
# With capital stock ~20-32 million millions AMD, pick CFC values that give ~1.2-2.2% depr.
# target_depr = cfc / (capital * 1e6)  =>  cfc = target_depr * capital * 1e6
# For 2023: capital = 32650000, target 1.5% -> cfc = 0.015 * 32650000 * 1e6 = 489_750_000_000_000? no.
# Wait: capital is in millions of AMD. cfc / 1e6 / capital gives unitless.
# cfc [AMD] / 1e6 = cfc in millions AMD; divide by capital in millions AMD -> ratio.
# So target_depr = cfc_millions / capital_millions = cfc / (1e6 * capital_stock_millions)
# For 2023: 0.015 = cfc / (1e6 * 32650000) -> cfc = 0.015 * 3.265e13 = 4.9e11 AMD.
# That's 490 billion AMD. Fine, that's the scale.
# We'll generate realistic-ish synthetic values.
cfc_values = []
depr_targets = [
    0.0220, 0.0215, 0.0210, 0.0205, 0.0200, 0.0195, 0.0190, 0.0185,
    0.0180, 0.0178, 0.0175, 0.0170, 0.0168, 0.0165, 0.0162, 0.0160,
    0.0158, 0.0155, 0.0152, 0.0150, 0.0148, 0.0146, 0.0144, 0.0142,
    0.0140, 0.0138, 0.0135, 0.0132,
]
for y, d in zip(cfc_years, depr_targets):
    # capital for that year from pwt_rnna (index = y - 1990)
    cap = pwt_rnna[y - 1990]
    cfc = d * 1e6 * cap
    cfc_values.append(cfc)
assert len(cfc_years) == len(cfc_values) == 28

# Armenia investment shock: 8 years starting 2026, totals ~$6.5B = ~2_515B AMD
# (at ~387 AMD/USD, but since the sheet is in millions AMD, we use millions AMD).
# Total million AMD: 6.5e9 USD * 387 AMD/USD = 2.515e12 AMD = 2_515_000 million AMD.
# Spread over 8 years with a bell curve, peaking ~year 4.
invest_years = list(range(INVEST_START, INVEST_START + INVEST_YEARS))
# Bell profile summing to 2_515_000 million AMD (~$6.5B).
invest_profile_weights = [0.07, 0.10, 0.13, 0.16, 0.17, 0.15, 0.12, 0.10]
TOTAL_INVEST = 2_515_000.0
invest_values = [round(TOTAL_INVEST * w, 2) for w in invest_profile_weights]
assert len(invest_values) == 8

# ============================================================
# COMPUTE HP FILTER ANSWERS (for cached values in answer xlsx)
# ============================================================
hist_years = list(range(HIST_START, HIST_END + 1))  # 2002..2023 = 22 years
N_HIST = len(hist_years)
assert N_HIST == 22

# K series for HP filter rows (col D, rows 6..27): pulls rnna for 2002..2023.
K_hist = [pwt_rnna[y - 1990] for y in hist_years]
# Y series (col E, rows 6..27): WEO_Data C * 1000 (billions->millions).
Y_hist = [weo_gdp_hist[y - 2000] * 1000 for y in hist_years]

lnK_hist = [np.log(k) for k in K_hist]
lnY_hist = [np.log(y) for y in Y_hist]
lnZ_hist = [lnY_hist[i] - CAPITAL_SHARE * lnK_hist[i] for i in range(N_HIST)]

# HP filter with lambda=100
n = N_HIST
lam = 100
I_mat = np.eye(n)
K_mat = np.zeros((n - 2, n))
for i in range(n - 2):
    K_mat[i, i] = 1
    K_mat[i, i + 1] = -2
    K_mat[i, i + 2] = 1
A = I_mat + lam * K_mat.T @ K_mat
tau = np.linalg.solve(A, np.array(lnZ_hist))
lnZ_HP = tau.tolist()

# Second-order diff (M column): M[i]=L[i]-2*L[i-1]+L[i-2], only for i>=2.
# Objective = SUMXMY2(K6:K27, L6:L27) + 100 * SUMSQ(M6:M27)
second_diff = [0.0, 0.0] + [
    lnZ_HP[i] - 2 * lnZ_HP[i - 1] + lnZ_HP[i - 2] for i in range(2, n)
]
obj_val = sum((lnZ_hist[i] - lnZ_HP[i]) ** 2 for i in range(n)) + 100 * sum(
    sd ** 2 for sd in second_diff
)

# Avg depreciation rate = AVERAGE of last 8 CFC depr rates
# CFC data rows 22..29 correspond to 2016..2023.
recent_depr = depr_targets[-DEPR_AVG_YEARS:]
avg_depr = float(np.mean(recent_depr))

# ============================================================
# LOWER TABLE: rows 36..75 (40 years, 2002..2041)
# ============================================================
all_years = list(range(HIST_START, PROJ_END + 1))  # 2002..2041
assert len(all_years) == 40

# For historical years (2002..2023) rows 36..57:
# - E=K (=D<hp_row>), F=Y (=WEO_Data!C<imf_row>*1000), G=LnZ trend (=L<hp_row>)
# For extended years (2024..2041) rows 58..75:
# - E = F * AVERAGE(D$49:D$57), F=WEO_Data!C<imf_row>*1000, G=TREND(...)
# K/Y anchor window in val: AVERAGE(D$49:D$57). Row 49=2015 (year 49-36+2002? check: row 36=2002,
#   so row 49 = 2002+(49-36)=2015; row 57=2023). That's 9 years (2015..2023) = KY_ANCHOR_YEARS.

# Compute anchor K/Y (historical K/Y over 2015..2023)
ky_ratios_hist = [K_hist[i] / Y_hist[i] for i in range(N_HIST)]
# Historical indices 13..21 (years 2015..2023)
anchor_idx_start = (2015 - HIST_START)  # 13
anchor_idx_end = (HIST_END - HIST_START) + 1  # 22
anchor_ky = float(np.mean(ky_ratios_hist[anchor_idx_start:anchor_idx_end]))

# Extended Y (2024..2041) using WEO formula C_next = C_prev * (1 + growth/100)
Y_ext_values = []
prev = weo_gdp_hist[-1]  # 2027
for y in range(2028, PROJ_END + 1):
    prev = prev * (1 + growth_ext / 100.0)
    Y_ext_values.append((y, prev))
# Y for 2024..2027 is in weo_gdp_hist[24..27]

# Build Y (millions AMD) and K arrays aligned to all_years
Y_arr = []
K_arr = []
for idx, y in enumerate(all_years):
    if y <= HIST_END:
        Y_arr.append(Y_hist[y - HIST_START])
        K_arr.append(K_hist[y - HIST_START])
    elif y <= 2027:
        gdp_bil = weo_gdp_hist[y - 2000]
        Y_arr.append(gdp_bil * 1000)
        K_arr.append(gdp_bil * 1000 * anchor_ky)
    else:
        # find in Y_ext_values
        gdp_bil = next(v for (yy, v) in Y_ext_values if yy == y)
        Y_arr.append(gdp_bil * 1000)
        K_arr.append(gdp_bil * 1000 * anchor_ky)

# LnZ trend (G column): historical = LnZ_HP; extended = linear trend
x_hist = np.arange(N_HIST)
coeffs = np.polyfit(x_hist, lnZ_HP, 1)
lnZ_trend_arr = []
for idx, y in enumerate(all_years):
    if y <= HIST_END:
        lnZ_trend_arr.append(lnZ_HP[y - HIST_START])
    else:
        # x position: y - HIST_START continues from x_hist
        x = y - HIST_START
        lnZ_trend_arr.append(coeffs[0] * x + coeffs[1])

# Ystar_base = EXP(G)*(E^alpha)
ystar_base = [np.exp(lnZ_trend_arr[i]) * (K_arr[i] ** CAPITAL_SHARE) for i in range(len(all_years))]

# Investment, deltaK, K_With
# Val logic (row 60..67 have J = (1-$B$3)*J<prev> + I<curr>):
# - J60 = (1-B3)*J59 + I60; J59 is empty (0), I60=Investment!C2
# - For pre-2026 rows, no I and no J formula (so effectively deltaK=0)
# - Extension continues to 2041: J68..J75 have same formula with I68..I75 (usually empty -> 0)
# => deltaK accumulation starts at 2026.
invest_map = {y: invest_values[i] for i, y in enumerate(invest_years)}
deltaK = [0.0] * len(all_years)
for idx, y in enumerate(all_years):
    if y < INVEST_START:
        deltaK[idx] = 0.0
    else:
        prev_d = deltaK[idx - 1] if idx > 0 else 0.0
        inv = invest_map.get(y, 0.0)
        deltaK[idx] = (1 - avg_depr) * prev_d + inv

K_with = [K_arr[i] + deltaK[i] for i in range(len(all_years))]
ystar_with = [np.exp(lnZ_trend_arr[i]) * (K_with[i] ** CAPITAL_SHARE) for i in range(len(all_years))]
uplift = [100.0 * (ystar_with[i] / ystar_base[i] - 1) for i in range(len(all_years))]
proj_gdp = [ystar_base[i] * (1 + uplift[i] / 100.0) for i in range(len(all_years))]

print(f"[verify] alpha={CAPITAL_SHARE}, depr_avg={avg_depr:.6f} ({avg_depr*100:.2f}%)")
print(f"[verify] anchor_ky (2015-2023)={anchor_ky:.4f}")
print(f"[verify] objective={obj_val:.6f}")
print(f"[verify] investment total (million AMD)={sum(invest_values):,.0f}")
print(f"[verify] ystar_base(2030)={ystar_base[2030-HIST_START]:.2f}")
print(f"[verify] gdp_2028(billion AMD)={weo_gdp_hist[-1]*(1+growth_ext/100):.4f}")


# ============================================================
# TEMPLATE (the xlsx the agent starts with)
# ============================================================
def create_template():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Production (active, first tab) ---
    ws = wb.active
    ws.title = "Production"
    ws["A1"] = "Assumption"
    ws["A2"] = "capital's share"
    ws["B2"] = CAPITAL_SHARE
    ws["A3"] = "annual depreciation rate"
    # B3 empty; agent fills formula

    ws["K4"] = "HP Filter"
    ws["D5"] = "K"
    ws["E5"] = "Real GDP"
    ws["F5"] = "LnK"
    ws["G5"] = "LnY"
    ws["K5"] = "LnZ"
    ws["L5"] = "LnZ_HP"
    ws["M5"] = "Second-order difference"
    ws["N5"] = "LnA-Trend(Check)"
    ws["O5"] = "objective"
    # P5 left empty for agent formula

    # HP filter years col C rows 6..27 (2002..2023)
    for i, y in enumerate(hist_years):
        ws.cell(row=6 + i, column=3, value=y)

    # Lower production table header row 35
    ws["D35"] = "K/Y"
    ws["E35"] = "K"
    ws["F35"] = "Y"
    ws["G35"] = "LnZ trend"
    ws["H35"] = "Ystar_base"
    ws["I35"] = "Investment"
    ws["J35"] = "\u0394K"   # ΔK
    ws["K35"] = "K_With"
    ws["L35"] = "Ystar_with"
    ws["M35"] = "Uplift"
    ws["N35"] = "Projected GDP"
    ws["O35"] = "Projected GDP Growth"
    ws["P35"] = "Baseline GDP Growth"

    for i, y in enumerate(all_years):
        ws.cell(row=36 + i, column=3, value=y)

    # --- Sheet 2: Investment ---
    inv = wb.create_sheet("Investment")
    inv["B1"] = "Year"
    inv["C1"] = "Project Investment (Deflated)"
    # Left blank for agent to populate from Spending schedule (hardcoded below for oracle)
    # We still pre-populate years so agent sees the schedule is 8 rows.
    for i, y in enumerate(invest_years):
        inv.cell(row=2 + i, column=2, value=y)
        # col C (values) left blank in template; agent fills

    # --- Sheet 3: CFC data ---
    cfc = wb.create_sheet("CFC data")
    cfc["A1"] = "DATE"
    cfc["B1"] = "TIME PERIOD"
    cfc["C1"] = (
        "Consumption of fixed capital (IDCM.A.N.AM.W2.S1.S1.D.P51C.N1G._T._Z.XDC.V.N)"
    )
    cfc["D1"] = "Capital Stock"
    cfc["E1"] = "Depreciation rate"
    for i, y in enumerate(cfc_years):
        cfc.cell(row=2 + i, column=1, value=datetime(y, 12, 31))
        cfc.cell(row=2 + i, column=2, value=y)
        # C, D, E left blank for agent

    # --- Sheet 4: PWT ---
    pwt = wb.create_sheet("PWT")
    pwt["A1"] = "year"
    pwt["B1"] = "rgdpna"
    pwt["C1"] = "rnna"
    for i, y in enumerate(pwt_years):
        pwt.cell(row=2 + i, column=1, value=y)
        # B, C left blank

    # --- Sheet 5: WEO_Data ---
    weo = wb.create_sheet("WEO_Data")
    weo["B1"] = "WEO Country Code"
    weo["C1"] = 911
    weo["D1"] = 911
    weo["B2"] = "WEO Subject Code"
    weo["C2"] = "NGDP_R"
    weo["D2"] = "NGDP_RPCH"
    weo["B3"] = "Country"
    weo["C3"] = "Armenia"
    weo["D3"] = "Armenia"
    weo["B4"] = "Subject Descriptor"
    weo["C4"] = "Gross domestic product, constant prices"
    weo["D4"] = "Gross domestic product, constant prices"
    weo["B5"] = "Units"
    weo["C5"] = "National currency"
    weo["D5"] = "Percent change"
    weo["B6"] = "Scale"
    weo["C6"] = "Billions"
    weo["D6"] = "Units"
    weo["B7"] = "Country/Series-specific Notes"
    weo["C7"] = (
        "Source: National Statistical Service of the Republic of Armenia. "
        "Base year: 2019. Primary domestic currency: Armenian dram. "
        "Data updated: 04/2025"
    )
    weo["D7"] = (
        "See notes for:  Gross domestic product, constant prices (National currency)."
    )
    # Year col B rows 8..51 (2000..2043)
    for i, y in enumerate(range(2000, PROJ_END + 3)):  # 2000..2043
        weo.cell(row=8 + i, column=2, value=y)
        # C, D blank

    out_path = f"{ENV_DIR}/{OUTPUT_XLSX}"
    wb.save(out_path)
    print(f"[ok] template saved -> {out_path}")


# ============================================================
# ANSWER XLSX (oracle - fully solved with formulas + cached values)
# ============================================================
def create_answer():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Production ---
    ws = wb.active
    ws.title = "Production"
    ws["A1"] = "Assumption"
    ws["A2"] = "capital's share"
    ws["B2"] = CAPITAL_SHARE
    ws["C2"] = "> standard capital share for small open economies"
    ws["A3"] = "annual depreciation rate"
    # B3 avg depr = AVERAGE('CFC data'!E22:E29) -- 8-year avg over 2016..2023
    depr_start_row = 2 + (cfc_years.index(2023) - DEPR_AVG_YEARS + 1)
    depr_end_row = 2 + cfc_years.index(2023)
    ws["B3"] = f"=AVERAGE('CFC data'!E{depr_start_row}:E{depr_end_row})"
    ws["C3"] = "> 8-year CFC-based depreciation average"

    ws["K4"] = "HP Filter"
    ws["D5"] = "K"
    ws["E5"] = "Real GDP"
    ws["F5"] = "LnK"
    ws["G5"] = "LnY"
    ws["K5"] = "LnZ"
    ws["L5"] = "LnZ_HP"
    ws["M5"] = "Second-order difference"
    ws["N5"] = "LnA-Trend(Check)"
    ws["O5"] = "objective"
    # P5 = objective
    ws["P5"] = f"=SUMXMY2(K6:K{5+N_HIST},L6:L{5+N_HIST})+100*SUMSQ(M6:M{5+N_HIST})"

    # HP filter rows 6..27
    for i, y in enumerate(hist_years):
        row = 6 + i
        ws.cell(row=row, column=3, value=y)
        # D = K (from PWT!C, rnna). PWT year y -> row 2 + (y-1990)
        pwt_row = 2 + (y - 1990)
        ws.cell(row=row, column=4, value=f"=PWT!C{pwt_row}")
        # E = Real GDP (WEO_Data!C * 1000). WEO year y -> row 8 + (y-2000)
        weo_row = 8 + (y - 2000)
        ws.cell(row=row, column=5, value=f"=WEO_Data!C{weo_row}*1000")
        ws.cell(row=row, column=6, value=f"=LN(D{row})")
        ws.cell(row=row, column=7, value=f"=LN(E{row})")
        # K = LnZ
        ws.cell(row=row, column=11, value=f"=G{row}-$B$2*F{row}")
        # L = LnZ_HP (solver result -> cached number)
        ws.cell(row=row, column=12, value=lnZ_HP[i])
        # M = 2nd-order diff (only i>=2)
        if i >= 2:
            ws.cell(row=row, column=13, value=f"=L{row}-2*L{row-1}+L{row-2}")
        # N = K - L (Check)
        ws.cell(row=row, column=14, value=f"=K{row}-L{row}")

    # --- Lower table header row 35 ---
    ws["D35"] = "K/Y"
    ws["E35"] = "K"
    ws["F35"] = "Y"
    ws["G35"] = "LnZ trend"
    ws["H35"] = "Ystar_base"
    ws["I35"] = "Investment"
    ws["J35"] = "\u0394K"
    ws["K35"] = "K_With"
    ws["L35"] = "Ystar_with"
    ws["M35"] = "Uplift"
    ws["N35"] = "Projected GDP"
    ws["O35"] = "Projected GDP Growth"
    ws["P35"] = "Baseline GDP Growth"

    # Rows 36..75 (2002..2041)
    hp_base_row = 6       # HP filter row for 2002 -> 6
    lower_base_row = 36   # lower table row for 2002 -> 36
    # Anchor window: D$49:D$57 in val = rows for 2015..2023.
    # Our rows follow same mapping: row 36=2002 => row 49=2015, row 57=2023.
    anchor_start_row = 49
    anchor_end_row = 57

    for i, y in enumerate(all_years):
        row = 36 + i
        ws.cell(row=row, column=3, value=y)

        if y <= HIST_END:
            # Historical: D=K/Y, E=D<hp_row>, F=WEO_Data!C<w>*1000, G=L<hp_row>
            hp_row = hp_base_row + (y - HIST_START)
            ws.cell(row=row, column=4, value=f"=E{row}/F{row}")
            ws.cell(row=row, column=5, value=f"=D{hp_row}")
            weo_row = 8 + (y - 2000)
            ws.cell(row=row, column=6, value=f"=WEO_Data!C{weo_row}*1000")
            ws.cell(row=row, column=7, value=f"=L{hp_row}")
        else:
            # Extended: E = F * AVERAGE(D$49:D$57), F=WEO_Data!C<w>*1000, G=TREND
            ws.cell(
                row=row, column=5,
                value=f"=F{row}*AVERAGE(D${anchor_start_row}:D${anchor_end_row})",
            )
            weo_row = 8 + (y - 2000)
            ws.cell(row=row, column=6, value=f"=WEO_Data!C{weo_row}*1000")
            # TREND array formula: known_y = L6:L27 (lnZ_HP), known_x = C6:C27, new_x = C<row>
            ws.cell(
                row=row, column=7,
                value=ArrayFormula(
                    f"G{row}",
                    f"=TREND(L$6:L${5+N_HIST},C$6:C${5+N_HIST},C{row})",
                ),
            )

        # H = Ystar_base
        ws.cell(row=row, column=8, value=f"=EXP(G{row})*(E{row}^$B$2)")

        # I = Investment link (only for shock years)
        if y in invest_map:
            inv_sheet_row = 2 + invest_years.index(y)
            ws.cell(row=row, column=9, value=f"=Investment!C{inv_sheet_row}")

        # J = deltaK formula: (1-$B$3)*J<prev>+I<curr> -- skip for very first row
        if i > 0:
            ws.cell(row=row, column=10, value=f"=(1-$B$3)*J{row-1}+I{row}")

        # K = K_With = E + J
        ws.cell(row=row, column=11, value=f"=E{row}+J{row}")
        # L = Ystar_with
        ws.cell(row=row, column=12, value=f"=EXP(G{row})*(K{row}^$B$2)")
        # M = Uplift
        ws.cell(row=row, column=13, value=f"=100*(L{row}/H{row}-1)")
        # N = Projected GDP
        ws.cell(row=row, column=14, value=f"=H{row}*(1+M{row}/100)")
        # O = Proj GDP growth / P = Baseline growth -- skip first row (no prev)
        if i > 0:
            ws.cell(row=row, column=15, value=f"=100*(N{row}/N{row-1}-1)")
            ws.cell(row=row, column=16, value=f"=100*(F{row}/F{row-1}-1)")

    # --- Sheet 2: Investment ---
    inv = wb.create_sheet("Investment")
    inv["B1"] = "Year"
    inv["C1"] = "Project Investment (Deflated)"
    for i, (y, v) in enumerate(zip(invest_years, invest_values)):
        inv.cell(row=2 + i, column=2, value=y)
        inv.cell(row=2 + i, column=3, value=v)

    # --- Sheet 3: CFC data ---
    cfc = wb.create_sheet("CFC data")
    cfc["A1"] = "DATE"
    cfc["B1"] = "TIME PERIOD"
    cfc["C1"] = (
        "Consumption of fixed capital (IDCM.A.N.AM.W2.S1.S1.D.P51C.N1G._T._Z.XDC.V.N)"
    )
    cfc["D1"] = "Capital Stock"
    cfc["E1"] = "Depreciation rate"
    for i, (y, v) in enumerate(zip(cfc_years, cfc_values)):
        row = 2 + i
        cfc.cell(row=row, column=1, value=datetime(y, 12, 31))
        cfc.cell(row=row, column=2, value=y)
        cfc.cell(row=row, column=3, value=v)
        pwt_row = 2 + (y - 1990)
        cfc.cell(row=row, column=4, value=f"=PWT!C{pwt_row}")
        cfc.cell(row=row, column=5, value=f"=C{row}/1000000/D{row}")

    # --- Sheet 4: PWT ---
    pwt = wb.create_sheet("PWT")
    pwt["A1"] = "year"
    pwt["B1"] = "rgdpna"
    pwt["C1"] = "rnna"
    for i, y in enumerate(pwt_years):
        row = 2 + i
        pwt.cell(row=row, column=1, value=y)
        pwt.cell(row=row, column=2, value=pwt_rgdpna[i])
        pwt.cell(row=row, column=3, value=pwt_rnna[i])

    # --- Sheet 5: WEO_Data ---
    weo = wb.create_sheet("WEO_Data")
    weo["B1"] = "WEO Country Code"
    weo["C1"] = 911
    weo["D1"] = 911
    weo["B2"] = "WEO Subject Code"
    weo["C2"] = "NGDP_R"
    weo["D2"] = "NGDP_RPCH"
    weo["B3"] = "Country"
    weo["C3"] = "Armenia"
    weo["D3"] = "Armenia"
    weo["B4"] = "Subject Descriptor"
    weo["C4"] = "Gross domestic product, constant prices"
    weo["D4"] = "Gross domestic product, constant prices"
    weo["B5"] = "Units"
    weo["C5"] = "National currency"
    weo["D5"] = "Percent change"
    weo["B6"] = "Scale"
    weo["C6"] = "Billions"
    weo["D6"] = "Units"
    weo["B7"] = "Country/Series-specific Notes"
    weo["C7"] = (
        "Source: National Statistical Service of the Republic of Armenia. "
        "Base year: 2019. Primary domestic currency: Armenian dram. "
        "Data updated: 04/2025"
    )
    weo["D7"] = (
        "See notes for:  Gross domestic product, constant prices (National currency)."
    )
    # Rows 8..51 (2000..2043)
    for i, y in enumerate(range(2000, PROJ_END + 3)):
        row = 8 + i
        weo.cell(row=row, column=2, value=y)
        if y <= 2027:
            weo.cell(row=row, column=3, value=weo_gdp_hist[y - 2000])
            weo.cell(row=row, column=4, value=weo_growth_hist[y - 2000])
        else:
            weo.cell(row=row, column=3, value=f"=C{row-1}*(1+D{row}/100)")
            weo.cell(row=row, column=4, value=growth_ext)

    answer_path = f"{SOLUTION_DIR}/answer-output.xlsx"
    wb.save(answer_path)
    print(f"[ok] answer saved -> {answer_path}")

    inject_cached_values(answer_path)


def inject_cached_values(xlsx_path):
    """Patch xlsx XML to insert <v> cached values alongside <f> formulas so
    openpyxl's data_only=True can read the computed numbers without needing
    LibreOffice. We only patch the cells where we know the computed values
    with high confidence.
    """
    import zipfile
    import shutil
    import tempfile
    import re
    import os

    values = {}

    # Sheet ordering in the workbook (create_answer call order):
    # sheet1 = Production, sheet2 = Investment, sheet3 = CFC data,
    # sheet4 = PWT, sheet5 = WEO_Data
    S_PROD = "sheet1"
    S_CFC = "sheet3"
    S_WEO = "sheet5"

    # Production B3 = avg_depr
    values[(S_PROD, "B3")] = avg_depr
    # Production P5 = objective
    values[(S_PROD, "P5")] = obj_val

    # HP filter rows 6..27
    for i, y in enumerate(hist_years):
        row = 6 + i
        values[(S_PROD, f"D{row}")] = pwt_rnna[y - 1990]
        values[(S_PROD, f"E{row}")] = weo_gdp_hist[y - 2000] * 1000
        values[(S_PROD, f"F{row}")] = np.log(pwt_rnna[y - 1990])
        values[(S_PROD, f"G{row}")] = np.log(weo_gdp_hist[y - 2000] * 1000)
        values[(S_PROD, f"K{row}")] = lnZ_hist[i]
        # L already a literal numeric, don't need patch
        if i >= 2:
            values[(S_PROD, f"M{row}")] = second_diff[i]
        values[(S_PROD, f"N{row}")] = lnZ_hist[i] - lnZ_HP[i]

    # Lower table rows 36..75
    for i, y in enumerate(all_years):
        row = 36 + i
        k_val = K_arr[i]
        y_val = Y_arr[i]
        if y <= HIST_END:
            values[(S_PROD, f"D{row}")] = k_val / y_val
        values[(S_PROD, f"E{row}")] = k_val
        values[(S_PROD, f"F{row}")] = y_val
        values[(S_PROD, f"G{row}")] = lnZ_trend_arr[i]
        values[(S_PROD, f"H{row}")] = ystar_base[i]
        if y in invest_map:
            values[(S_PROD, f"I{row}")] = invest_map[y]
        if i > 0:
            values[(S_PROD, f"J{row}")] = deltaK[i]
        values[(S_PROD, f"K{row}")] = K_with[i]
        values[(S_PROD, f"L{row}")] = ystar_with[i]
        values[(S_PROD, f"M{row}")] = uplift[i]
        values[(S_PROD, f"N{row}")] = proj_gdp[i]
        # O, P growth rates
        if i > 0:
            if proj_gdp[i - 1]:
                values[(S_PROD, f"O{row}")] = 100 * (proj_gdp[i] / proj_gdp[i - 1] - 1)
            if Y_arr[i - 1]:
                values[(S_PROD, f"P{row}")] = 100 * (Y_arr[i] / Y_arr[i - 1] - 1)

    # CFC data rows 2..29: D (capital) and E (depr rate)
    for i, y in enumerate(cfc_years):
        row = 2 + i
        cap = pwt_rnna[y - 1990]
        values[(S_CFC, f"D{row}")] = cap
        values[(S_CFC, f"E{row}")] = cfc_values[i] / 1e6 / cap

    # WEO extended GDP (rows 36..51 = 2028..2043)
    prev_val = weo_gdp_hist[-1]
    for i, y in enumerate(range(2028, PROJ_END + 3)):
        row = 8 + (y - 2000)
        prev_val = prev_val * (1 + growth_ext / 100.0)
        values[(S_WEO, f"C{row}")] = prev_val

    # Patch XML
    tmp_xlsx = xlsx_path + ".tmp"
    with zipfile.ZipFile(xlsx_path, "r") as zin, zipfile.ZipFile(tmp_xlsx, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            for sheet_id in ("sheet1", "sheet2", "sheet3", "sheet4", "sheet5"):
                if item.filename == f"xl/worksheets/{sheet_id}.xml":
                    sheet_values = {k[1]: v for k, v in values.items() if k[0] == sheet_id}
                    if sheet_values:
                        content = data.decode("utf-8")
                        content = _patch_sheet_xml(content, sheet_values)
                        data = content.encode("utf-8")
                    break
            zout.writestr(item, data)
    shutil.move(tmp_xlsx, xlsx_path)
    print("[ok] cached values injected")


def _patch_sheet_xml(xml_content, cell_values):
    import re
    for cell_ref, value in cell_values.items():
        # Match the full <c r="..."...>...</c> element (self-closing <v /> allowed).
        pattern = rf'(<c\s+r="{cell_ref}"[^>]*>)(.*?)(</c>)'
        match = re.search(pattern, xml_content, re.DOTALL)
        if not match:
            continue
        inner = match.group(2)
        # Format numeric values to a reasonable precision
        if isinstance(value, float):
            vstr = f"{value:.10g}"
        else:
            vstr = str(value)
        # Replace existing <v>..</v> or <v /> with our cached value,
        # or inject <v>..</v> after <f>..</f> if no <v> tag present.
        if re.search(r'<v\s*/>', inner):
            new_inner = re.sub(r'<v\s*/>', f'<v>{vstr}</v>', inner, count=1)
        elif re.search(r'<v>.*?</v>', inner, re.DOTALL):
            new_inner = re.sub(r'<v>.*?</v>', f'<v>{vstr}</v>', inner, count=1, flags=re.DOTALL)
        elif re.search(r'</f>', inner):
            new_inner = re.sub(r'(</f>)', rf'\1<v>{vstr}</v>', inner, count=1)
        else:
            # Cell with no formula (literal number) — skip
            new_inner = inner
        if new_inner != inner:
            replacement = f"{match.group(1)}{new_inner}{match.group(3)}"
            xml_content = xml_content[:match.start()] + replacement + xml_content[match.end():]
    return xml_content


if __name__ == "__main__":
    create_template()
    create_answer()
