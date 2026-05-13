"""
Unit tests for export shock analysis task (Armenia, supply side).
Verifies that the Excel file contains proper formulas and correct calculations.
"""
import pytest
import openpyxl
from pathlib import Path

# The test file that the agent should modify
TEST_FILE = Path("/root/test - export.xlsx")

# Expected sheets in the workbook
REQUIRED_SHEETS = ["WEO_Projections", "Impact", "EXPORTS-2023", "IMPORTS-2023", "Trade_Calc"]

# Tolerance for floating point comparisons
TOLERANCE = 0.01


def load_workbook_with_formulas():
    """Load workbook preserving formulas."""
    assert TEST_FILE.exists(), f"Test file not found: {TEST_FILE}"
    return openpyxl.load_workbook(TEST_FILE)


def load_workbook_with_values():
    """Load workbook with calculated values."""
    assert TEST_FILE.exists(), f"Test file not found: {TEST_FILE}"
    return openpyxl.load_workbook(TEST_FILE, data_only=True)


def test_required_sheets_exist():
    """Test that all required sheets exist in the workbook."""
    wb = load_workbook_with_formulas()
    sheet_names = [s.lower().strip() for s in wb.sheetnames]

    for required in REQUIRED_SHEETS:
        # Case-insensitive and whitespace-tolerant matching
        required_lower = required.lower().strip()
        found = any(required_lower == s for s in sheet_names)
        assert found, f"Required sheet '{required}' not found. Available: {wb.sheetnames}"


def get_sheet_case_insensitive(wb, name):
    """Get sheet by name with case-insensitive matching."""
    name_lower = name.lower().strip()
    for sheet_name in wb.sheetnames:
        if sheet_name.lower().strip() == name_lower:
            return wb[sheet_name]
    raise KeyError(f"Sheet '{name}' not found")


def test_weo_projections_has_formulas():
    """Test that WEO_Projections sheet has formulas for projected years (not hardcoded)."""
    wb = load_workbook_with_formulas()
    weo = get_sheet_case_insensitive(wb, "WEO_Projections")

    # Check that extended years (rows 8+ = 2027+) have formulas for Nominal GDP (col B)
    for row in range(8, 13):
        cell_b = weo.cell(row=row, column=2)  # Column B - Nominal GDP
        cell_value = cell_b.value
        if cell_value is not None:
            is_formula = isinstance(cell_value, str) and cell_value.startswith("=")
            assert is_formula, f"Cell B{row} should contain a formula for projected GDP, got: {cell_value}"

    # Check GDP deflator YoY change has formulas (col E, rows 3-7)
    for row in range(3, 8):
        cell_e = weo.cell(row=row, column=5)  # Column E - Deflator YoY change
        cell_value = cell_e.value
        if cell_value is not None:
            is_formula = isinstance(cell_value, str) and cell_value.startswith("=")
            assert is_formula, f"Cell E{row} should contain a formula for deflator change, got: {cell_value}"


def test_trade_calc_formulas_and_export_ratio():
    """Test that Trade_Calc sheet has proper formulas linking to EXPORTS/IMPORTS and calculates export dependency ratio."""
    wb_formulas = load_workbook_with_formulas()
    wb_values = load_workbook_with_values()

    tc_formulas = get_sheet_case_insensitive(wb_formulas, "Trade_Calc")
    tc_values = get_sheet_case_insensitive(wb_values, "Trade_Calc")

    # Check that B4 links to EXPORTS sheet
    b4_value = tc_formulas["B4"].value
    assert b4_value is not None, "Cell B4 should have a value"
    is_formula = isinstance(b4_value, str) and b4_value.startswith("=")
    if is_formula:
        assert "EXPORT" in b4_value.upper(), f"B4 formula should reference EXPORTS sheet, got: {b4_value}"

    # Check that C4 links to IMPORTS sheet
    c4_value = tc_formulas["C4"].value
    if c4_value is not None:
        is_formula = isinstance(c4_value, str) and c4_value.startswith("=")
        if is_formula:
            assert "IMPORT" in c4_value.upper(), f"C4 formula should reference IMPORTS sheet, got: {c4_value}"

    # Check export dependency ratio in C43
    c43_formula = tc_formulas["C43"].value
    assert c43_formula is not None, "Cell C43 (export dependency ratio) should have a value"
    is_formula = isinstance(c43_formula, str) and c43_formula.startswith("=")
    assert is_formula, f"C43 should contain a formula for export dependency ratio, got: {c43_formula}"

    # Verify the calculated export dependency ratio is reasonable (between 0 and 1)
    c43_value = tc_values["C43"].value
    if c43_value is not None:
        assert 0 < c43_value < 1, f"Export dependency ratio should be between 0 and 1, got: {c43_value}"


@pytest.mark.parametrize("scenario,expected_multiplier,expected_export_ratio", [
    (1, 0.7, None),      # Scenario 1: multiplier=0.7, export ratio from Trade_Calc
    (2, 0.9, None),      # Scenario 2: multiplier=0.9
    (3, 0.7, 0.4),       # Scenario 3: export dependency ratio=0.4
])
def test_impact_scenarios_assumptions(scenario, expected_multiplier, expected_export_ratio):
    """Test that Impact sheet has correct assumptions for each scenario."""
    wb_values = load_workbook_with_values()
    wb_formulas = load_workbook_with_formulas()

    impact_values = get_sheet_case_insensitive(wb_values, "Impact")
    impact_formulas = get_sheet_case_insensitive(wb_formulas, "Impact")

    # Find scenario sections by searching for "Scenario" text
    scenario_row = None
    for row in range(1, 60):
        for col in range(1, 10):
            cell = impact_values.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                if f"scenario {scenario}" in cell.value.lower() or (scenario == 1 and "assumptions" in cell.value.lower() and row < 20):
                    scenario_row = row
                    break
        if scenario_row:
            break

    if scenario == 1:
        # Scenario 1 assumptions at rows 13-16
        # Check total shock formula (4200 * 387.5)
        d13 = impact_values["D13"].value
        expected_shock = 4200 * 387.5
        if d13 is not None:
            assert abs(d13 - expected_shock) < TOLERANCE * expected_shock, \
                f"Scenario 1 total export shock should be ~{expected_shock}, got {d13}"

        # Check supply multiplier
        d15 = impact_values["D15"].value
        if d15 is not None:
            assert abs(d15 - expected_multiplier) < TOLERANCE, \
                f"Scenario 1 supply multiplier should be {expected_multiplier}, got {d15}"

        # Check export dep ratio links to Trade_Calc (should be a formula)
        d14_formula = impact_formulas["D14"].value
        if d14_formula is not None and isinstance(d14_formula, str):
            assert "TRADE" in d14_formula.upper() or d14_formula.startswith("="), \
                f"Scenario 1 export dependency ratio should link to Trade_Calc sheet, got: {d14_formula}"
    else:
        # For scenarios 2 and 3, find the assumptions section
        if scenario_row is not None:
            found_correct_multiplier = False
            found_correct_export_ratio = expected_export_ratio is None

            for check_row in range(scenario_row, scenario_row + 15):
                for col in range(1, 10):
                    cell_val = impact_values.cell(row=check_row, column=col).value
                    if cell_val is not None:
                        if isinstance(cell_val, (int, float)):
                            if expected_multiplier is not None and abs(cell_val - expected_multiplier) < TOLERANCE:
                                found_correct_multiplier = True
                            if expected_export_ratio is not None and abs(cell_val - expected_export_ratio) < TOLERANCE:
                                found_correct_export_ratio = True

            if scenario == 2:
                assert found_correct_multiplier, \
                    f"Scenario 2 should have supply multiplier = {expected_multiplier}"
            if scenario == 3 and expected_export_ratio is not None:
                assert found_correct_export_ratio, \
                    f"Scenario 3 should have export dependency ratio = {expected_export_ratio}"


def test_impact_project_allocation_bell_shape():
    """Test that project allocation follows bell shape pattern (0.08-0.17-0.25-0.25-0.17-0.08)."""
    wb_values = load_workbook_with_values()
    impact = get_sheet_case_insensitive(wb_values, "Impact")

    expected_allocation = [0.08, 0.17, 0.25, 0.25, 0.17, 0.08]

    # Find project allocation column (column D, rows 4-9 for 6 years 2027-2032)
    allocation_values = []
    for row in range(4, 10):  # 6 years of allocation
        cell_d = impact.cell(row=row, column=4)  # Column D
        if cell_d.value is not None and isinstance(cell_d.value, (int, float)):
            allocation_values.append(cell_d.value)

    if len(allocation_values) >= 6:
        for i, (expected, actual) in enumerate(zip(expected_allocation, allocation_values[:6])):
            assert abs(expected - actual) < TOLERANCE, \
                f"Project allocation year {i+1} should be {expected}, got {actual}"
