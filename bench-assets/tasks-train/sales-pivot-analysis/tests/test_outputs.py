#!/usr/bin/env python3
"""
Tests for Canadian Education Pivot Table Analysis task.
Validates that the output file contains correctly structured pivot tables
with the correct row/column fields and aggregation types.
"""

import pytest
from openpyxl import load_workbook

OUTPUT_FILE = "/root/education_report.xlsx"
PERFORMANCE_XLSX = "/root/performance.xlsx"

# Sheet configurations: (sheet_name, expected_aggregation, col_field_name or None)
PIVOT_SHEETS = [
    ("Enrollment by Province", "sum", None),
    ("Teachers by Province", "sum", None),
    ("Schools by Province", "count", None),
    ("Province Funding Tier", "sum", "tier"),
]

# Required columns in SourceData: (description, match_fn)
REQUIRED_COLUMNS = [
    ("SCHOOL_ID", lambda h: "school_id" in h or "schoolid" in h.replace("_", "")),
    ("SCHOOL_NAME", lambda h: "school_name" in h or "schoolname" in h.replace("_", "")),
    ("PROVINCE", lambda h: "province" in h),
    ("ENROLLMENT_2024", lambda h: "enrollment" in h),
    ("TEACHERS", lambda h: "teachers" in h),
    ("AVG_SCORE", lambda h: "avg" in h and "score" in h),
    ("FUNDING", lambda h: "funding" in h),
    ("Tier", lambda h: "tier" in h),
    ("Budget", lambda h: h == "budget"),
]


@pytest.fixture(scope="module")
def workbook():
    """Load output workbook once for all tests."""
    return load_workbook(OUTPUT_FILE)


def _get_pivot_field_names(pivot):
    """Extract field names from pivot table cache."""
    cache = pivot.cache
    if cache and cache.cacheFields:
        return [f.name for f in cache.cacheFields]
    return []


def _get_field_name_by_index(pivot, fields):
    """Get field name from pivot fields collection."""
    field_names = _get_pivot_field_names(pivot)
    if fields and len(fields) > 0:
        idx = fields[0].x
        if idx is not None and 0 <= idx < len(field_names):
            return field_names[idx]
    return None


class TestPivotTableConfiguration:
    """Test pivot tables have correct row field, aggregation, and column field (if matrix)."""

    @pytest.mark.parametrize("sheet_name,expected_agg,col_field", PIVOT_SHEETS)
    def test_pivot_row_is_province(self, workbook, sheet_name, expected_agg, col_field):
        """Pivot row field should be PROVINCE."""
        pivot = workbook[sheet_name]._pivots[0]
        row_field = _get_field_name_by_index(pivot, pivot.rowFields)
        assert row_field and "province" in row_field.lower(), f"Row field should be PROVINCE, got '{row_field}'"

    @pytest.mark.parametrize("sheet_name,expected_agg,col_field", PIVOT_SHEETS)
    def test_pivot_uses_correct_aggregation(self, workbook, sheet_name, expected_agg, col_field):
        """Pivot data field should use correct aggregation."""
        pivot = workbook[sheet_name]._pivots[0]
        data_field = pivot.dataFields[0]
        assert data_field.subtotal == expected_agg, f"Expected '{expected_agg}' aggregation, got '{data_field.subtotal}'"

    @pytest.mark.parametrize("sheet_name,expected_agg,col_field", PIVOT_SHEETS)
    def test_pivot_col_field(self, workbook, sheet_name, expected_agg, col_field):
        """Matrix pivots must have column fields configured correctly."""
        if not col_field:
            pytest.skip(f"'{sheet_name}' is not a matrix pivot")
        pivot = workbook[sheet_name]._pivots[0]
        actual_col = _get_field_name_by_index(pivot, pivot.colFields)
        assert actual_col and col_field in actual_col.lower(), f"Column field should be '{col_field}', got '{actual_col}'"


@pytest.fixture(scope="module")
def source_sheet(workbook):
    """Find the source data sheet."""
    for name in workbook.sheetnames:
        if "source" in name.lower() or "data" in name.lower():
            return workbook[name]
    pytest.fail("No source data sheet found (expected sheet name containing 'source' or 'data')")


@pytest.fixture(scope="module")
def headers(source_sheet):
    """Get headers from source sheet."""
    first_row = next(source_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    return [str(h).strip().lower() if h else "" for h in first_row]


class TestSourceDataSheet:
    """Test that the SourceData sheet has required columns."""

    @pytest.mark.parametrize("desc,match_fn", REQUIRED_COLUMNS)
    def test_source_data_has_required_column(self, headers, desc, match_fn):
        """SourceData must have required column."""
        assert any(match_fn(h) for h in headers), f"Missing {desc} column. Found: {headers}"


VALID_TIERS = {"T1", "T2", "T3", "T4"}
VALID_PROVINCES = {
    "Ontario", "Quebec", "British Columbia", "Alberta",
    "Manitoba", "Saskatchewan", "Nova Scotia",
    "New Brunswick", "Newfoundland and Labrador", "Prince Edward Island"
}


@pytest.fixture(scope="module")
def source_data(source_sheet):
    """Parse source data into list of dicts."""
    rows = list(source_sheet.iter_rows(values_only=True))
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    data = [dict(zip(headers, row)) for row in rows[1:] if any(cell is not None for cell in row)]
    return data, headers


@pytest.fixture(scope="module")
def performance_data():
    """Parse input performance data."""
    wb = load_workbook(PERFORMANCE_XLSX)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    data = [dict(zip(headers, row)) for row in rows[1:] if any(cell is not None for cell in row)]
    return data, headers


class TestSourceDataContent:
    """Test that SourceData contains correct data."""

    def test_source_data_has_reasonable_row_count(self, source_data):
        """SourceData must have joined data (approximately 700-900 schools)."""
        data, _ = source_data
        assert 600 <= len(data) <= 1000, f"Expected 600-1000 schools, got {len(data)}"

    def test_tier_values_are_valid(self, source_data):
        """Tier column must contain valid T1-T4 values."""
        data, headers = source_data
        tier_col = next((h for h in headers if "tier" in h.lower()), None)
        tiers_found = {row.get(tier_col) for row in data if row.get(tier_col)}
        invalid = tiers_found - VALID_TIERS
        assert not invalid, f"Invalid tier values: {invalid}"

    def test_province_values_are_valid(self, source_data):
        """PROVINCE column must contain valid Canadian provinces."""
        data, headers = source_data
        province_col = next((h for h in headers if "province" in h.lower()), None)
        provinces_found = {row.get(province_col) for row in data if row.get(province_col)}
        invalid = provinces_found - VALID_PROVINCES
        assert not invalid, f"Invalid provinces: {invalid}"


class TestDataTransformationCorrectness:
    """Test data transformation correctness (anti-cheating)."""

    def test_pivot_cache_has_fields(self, workbook):
        """Pivot cache must have field definitions."""
        pivot = workbook["Enrollment by Province"]._pivots[0]
        assert len(pivot.cache.cacheFields) > 0, "Pivot cache has no field definitions"

    def test_budget_equals_teachers_times_funding(self, source_data):
        """Budget must equal TEACHERS × FUNDING."""
        data, headers = source_data
        teachers_col = next((h for h in headers if "teachers" in h.lower()), None)
        funding_col = next((h for h in headers if "funding" in h.lower()), None)
        budget_col = next((h for h in headers if h.lower() == "budget"), None)

        errors = []
        for i, row in enumerate(data[:50]):
            teachers, funding, budget = row.get(teachers_col), row.get(funding_col), row.get(budget_col)
            if all(v is not None for v in (teachers, funding, budget)):
                try:
                    if abs(float(teachers) * float(funding) - float(budget)) > 1:
                        errors.append(f"Row {i+2}: {teachers}×{funding}≠{budget}")
                except (ValueError, TypeError):
                    pass
        assert not errors, f"Budget calculation errors:\n" + "\n".join(errors[:5])

    def test_school_ids_from_performance_file_present(self, source_data, performance_data):
        """SCHOOL_IDs from performance file must be present in output (verifies join)."""
        out_data, out_headers = source_data
        in_data, in_headers = performance_data

        out_id_col = next((h for h in out_headers if "school" in h.lower() and "id" in h.lower()), None)
        in_id_col = next((h for h in in_headers if "school" in h.lower() and "id" in h.lower()), None)

        out_ids = {str(row.get(out_id_col)) for row in out_data if row.get(out_id_col)}
        in_ids = {str(row.get(in_id_col)) for row in in_data if row.get(in_id_col)}

        overlap = len(out_ids & in_ids)
        assert overlap > len(in_ids) * 0.9, f"Less than 90% SCHOOL_IDs found. {overlap}/{len(in_ids)}"
