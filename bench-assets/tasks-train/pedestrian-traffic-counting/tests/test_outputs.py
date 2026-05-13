"""Reduced verifier for the non-pedestrian transit-log variant.

We exercise a smaller subset of the original validation than the canonical
pedestrian-counting test:
  - existence of the report
  - workbook structure (single sheet, sheet name)
  - header / column schema
  - per-clip tolerance reward (1 / (1 + |diff|))
  - exact-match requirement on the integer counts

We deliberately drop the per-cell null/blank-handling check and the strict
row-by-row workbook equality check from the canonical suite; that lets the
reduced variant focus on the counting skill itself rather than on workbook
hygiene minutiae.
"""

import os
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPORT_FILE = Path("/app/video/transit_log.xlsx")
ORACLE_FILE = Path(__file__).resolve().parent / "transit_log.xlsx"
REWARD_FILE = Path("/logs/verifier/reward.txt")
SHEET = "tally"
COLUMNS = ["clip", "non_pedestrian_count"]


def _stringify(cell: Any) -> str:
    return "" if cell is None else str(cell)


def _grid(path: Path, sheet: str) -> list[list[str]]:
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[sheet]
        rows: list[list[str]] = []
        for r in range(1, (ws.max_row or 0) + 1):
            rows.append([_stringify(ws.cell(row=r, column=c).value)
                         for c in range(1, (ws.max_column or 0) + 1)])
        return rows
    finally:
        wb.close()


def _sheets(path: Path) -> list[str]:
    wb = load_workbook(path, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _to_pairs(rows: list[list[str]]) -> dict[str, str]:
    """Skip header, drop blank rows, return clip -> count_str."""
    out: dict[str, str] = {}
    for r in rows[1:]:
        if len(r) >= 2 and r[0].strip():
            out[r[0]] = r[1]
    return out


def _record_reward(value: float) -> None:
    REWARD_FILE.parent.mkdir(parents=True, exist_ok=True)
    REWARD_FILE.write_text(f"{value:.6f}\n")


def _per_clip_reward(actual: dict[str, str], oracle: dict[str, str]) -> float:
    if not oracle:
        return 0.0
    accum = 0.0
    for clip, want in oracle.items():
        got = actual.get(clip, "")
        try:
            want_n = float(want) if want.strip() else 0.0
        except ValueError:
            want_n = 0.0
        try:
            got_n = float(got) if got.strip() else 0.0
        except ValueError:
            got_n = 0.0
        diff = abs(got_n - want_n)
        clip_score = 1.0 / (1.0 + diff)
        marker = "OK " if diff == 0 else f"diff={diff:.1f}"
        print(f"  {clip}: got={got!r} want={want!r}  reward={clip_score:.4f} ({marker})")
        accum += clip_score
    mean = accum / len(oracle)
    print(f"\nMean per-clip reward: {mean:.4f}")
    return mean


def test_transit_log():
    """End-to-end check on the transit log workbook."""

    # 1) file must exist
    if not REPORT_FILE.exists():
        _record_reward(0.0)
        raise AssertionError(f"missing report: {REPORT_FILE}")

    try:
        # 2) one sheet, named correctly
        sheets = _sheets(REPORT_FILE)
        assert len(sheets) == 1, f"need exactly 1 sheet, found {sheets}"
        assert sheets[0] == SHEET, f"sheet must be {SHEET!r}, got {sheets[0]!r}"

        actual = _grid(REPORT_FILE, SHEET)
        oracle = _grid(ORACLE_FILE, SHEET)

        # 3) header schema
        assert actual, "report has no rows at all"
        assert actual[0] == COLUMNS, (
            f"header mismatch -> got {actual[0]}, want {COLUMNS}"
        )

        # 4) per-clip tolerance reward (always written)
        actual_pairs = _to_pairs(actual)
        oracle_pairs = _to_pairs(oracle)
        print("\n--- per-clip reward ---")
        reward = _per_clip_reward(actual_pairs, oracle_pairs)
        _record_reward(reward)

        # 5) exact-match assertion (gates dry-run pass)
        for clip, want in oracle_pairs.items():
            got = actual_pairs.get(clip, "<missing>")
            assert got == want, (
                f"clip {clip!r}: got {got!r}, expected {want!r}"
            )

    except Exception:
        # best-effort reward write on any failure path
        try:
            actual = _grid(REPORT_FILE, _sheets(REPORT_FILE)[0])
            oracle = _grid(ORACLE_FILE, SHEET)
            _record_reward(_per_clip_reward(_to_pairs(actual), _to_pairs(oracle)))
        except Exception:
            _record_reward(0.0)
        raise
