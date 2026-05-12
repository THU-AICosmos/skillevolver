"""
Tests for population-weighted savings rate calculation task.
Verifies that the Excel formulas are correctly implemented for:
- Step 1: Lookup formulas (INDEX/MATCH, VLOOKUP/MATCH, etc.) for data retrieval
- Step 2: Gross savings rate calculations and statistics
- Step 3: Population-weighted mean calculation using SUMPRODUCT
"""

import csv
import glob
from pathlib import Path

import pytest
from openpyxl import load_workbook

EXCEL_FILE = Path("/root/savings_rate.xlsx")
CSV_PATTERN = "/root/sheet.csv.*"
TOLERANCE = 0.5

_csv_data_cache = None
_task_sheet_index = None


def find_task_csv():
    """
    Locate the CSV file containing Task sheet data exported by ssconvert.

    Uses three strategies in order of reliability:
    1. Match by sheet index from xlsx (ssconvert names files sheet.csv.0, sheet.csv.1, etc.)
    2. Content pattern matching for Task sheet identifiers
    3. Fall back to first available CSV
    """
    global _task_sheet_index

    csv_files = sorted(glob.glob(CSV_PATTERN))
    if not csv_files:
        return None

    wb = load_workbook(EXCEL_FILE, data_only=False)
    for idx, name in enumerate(wb.sheetnames):
        if "Task" in name:
            _task_sheet_index = idx
            wb.close()
            expected_file = f"/root/sheet.csv.{idx}"
            if Path(expected_file).exists():
                return expected_file
            break
    wb.close()

    for csv_file in csv_files:
        try:
            with open(csv_file, encoding="utf-8", errors="ignore") as f:
                content = f.read(5000)
                if "GNI_CURR_USD" in content or "ASEAN" in content:
                    if "Savings Rate" in content or "Gross National Income" in content:
                        return csv_file
        except:
            continue

    return csv_files[0] if csv_files else None


def load_csv_data():
    """
    Load and cache CSV data with evaluated formula values.

    Returns a dict mapping cell references (e.g., 'F11') to their values,
    converting numeric strings to floats where possible.
    """
    global _csv_data_cache

    if _csv_data_cache is not None:
        return _csv_data_cache

    csv_file = find_task_csv()
    if csv_file is None:
        _csv_data_cache = {}
        return _csv_data_cache

    _csv_data_cache = {}
    try:
        with open(csv_file, encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, start=1):
                for col_idx, val in enumerate(row):
                    col_letter = chr(ord("A") + col_idx) if col_idx < 26 else None
                    if col_letter:
                        cell_ref = f"{col_letter}{row_idx}"
                        if val and val.strip():
                            try:
                                _csv_data_cache[cell_ref] = float(val)
                            except ValueError:
                                _csv_data_cache[cell_ref] = val
                        else:
                            _csv_data_cache[cell_ref] = None
    except Exception as e:
        print(f"Error loading CSV: {e}")
        _csv_data_cache = {}

    return _csv_data_cache


def get_workbook():
    """Load the workbook with data only (calculated values)."""
    return load_workbook(EXCEL_FILE, data_only=True)


def get_workbook_formulas():
    """Load the workbook with formulas."""
    return load_workbook(EXCEL_FILE, data_only=False)


def get_task_sheet(wb):
    """Find the Task sheet or return the active sheet."""
    for sheet_name in wb.sheetnames:
        if "Task" in sheet_name:
            return wb[sheet_name]
    return wb.active


def cell_value(ws, cell):
    """
    Get cell value, preferring xlsx direct values then falling back to CSV.

    CSV fallback is needed because openpyxl cannot evaluate formulas—ssconvert
    exports calculated values to CSV which we can read.
    """
    val = ws[cell].value
    if val is not None and isinstance(val, (int, float)):
        return val
    csv_val = cell_value_csv(cell)
    if csv_val is not None:
        return csv_val
    return val if val is not None else 0


def cell_value_csv(cell):
    """Get cell value from CSV only."""
    csv_data = load_csv_data()
    return csv_data.get(cell)


def is_formula(val):
    """Check if a value is a formula string."""
    return isinstance(val, str) and val.startswith("=")


def has_formula_or_value(ws, cell):
    """Check if cell has a formula or a value from either xlsx or CSV source."""
    formula_wb = get_workbook_formulas()
    formula_ws = get_task_sheet(formula_wb)
    xlsx_val = formula_ws[cell].value
    formula_wb.close()

    if xlsx_val is not None:
        return True

    csv_val = cell_value_csv(cell)
    if csv_val is not None:
        return True

    return False


# Expected values from the Task sheet answer key
EXPECTED_GNI = {
    11: [1050.3, 1148.7, 1289.4, 1370.2],
    12: [494.8, 516.3, 500.2, 530.1],
    13: [340.6, 368.5, 413.2, 450.8],
    14: [362.1, 394.5, 420.8, 440.3],
    15: [336.7, 372.4, 415.9, 430.5],
}

EXPECTED_CONS = {
    18: [805.2, 878.4, 982.1, 1058.6],
    19: [345.8, 365.2, 362.7, 391.5],
    20: [263.1, 280.4, 318.5, 352.7],
    21: [310.8, 335.2, 362.4, 381.9],
    22: [210.3, 233.8, 261.5, 278.2],
}

EXPECTED_POP = {
    25: [273.5, 276.4, 279.1, 281.6],
    26: [69.8, 69.9, 70.0, 70.1],
    27: [97.3, 98.2, 99.0, 99.7],
    28: [109.6, 111.0, 112.5, 113.9],
    29: [32.4, 32.8, 33.2, 33.6],
}

EXPECTED_SAVINGS_RATE = {
    33: [23.3, 23.5, 23.8, 22.7],
    34: [30.1, 29.3, 27.5, 26.1],
    35: [22.8, 23.9, 22.9, 21.8],
    36: [14.2, 15.0, 13.9, 13.3],
    37: [37.5, 37.2, 37.1, 35.4],
}

EXPECTED_STATS = {
    40: [14.2, 15.0, 13.9, 13.3],  # MIN
    41: [37.5, 37.2, 37.1, 35.4],  # MAX
    42: [23.3, 23.9, 23.8, 22.7],  # MEDIAN
    43: [25.6, 25.8, 25.0, 23.9],  # AVERAGE
    44: [22.8, 23.5, 22.9, 21.8],  # PERCENTILE(0.25)
    45: [30.1, 29.3, 27.5, 26.1],  # PERCENTILE(0.75)
    46: [8.7, 8.2, 8.4, 8.0],  # STDEV
}

# Weighted mean: SUMPRODUCT(savings_rate, population) / SUM(population)
EXPECTED_WEIGHTED_MEAN = [23.1, 23.4, 23.0, 21.9]

COLUMNS = ["F", "G", "H", "I"]


class TestFileExists:
    """Test that the Excel file exists and is readable."""

    def test_excel_file_exists(self):
        """Verify the Excel file exists."""
        assert EXCEL_FILE.exists(), f"Excel file not found at {EXCEL_FILE}"

    def test_excel_file_readable(self):
        """Verify the Excel file can be opened."""
        wb = get_workbook()
        assert wb is not None
        wb.close()

    def test_sheet_structure(self):
        """Debug test: show available data sources."""
        wb = get_workbook()
        ws = get_task_sheet(wb)
        print(f"\nAvailable sheets: {wb.sheetnames}")
        print(f"Task sheet: {ws.title}")

        csv_file = find_task_csv()
        print(f"CSV file: {csv_file}")

        csv_data = load_csv_data()
        print(f"CSV data entries: {len(csv_data)}")

        # Print sample values from CSV
        print("\nSample CSV values:")
        for cell in ["F11", "F33", "F40", "F49"]:
            print(f"  {cell}: {csv_data.get(cell)}")

        wb.close()
        assert True


class TestStep1LookupFormulas:
    """Test Step 1: Verify lookup formulas OR values exist for data retrieval."""

    def test_gni_have_data(self):
        """Verify GNI section (F11:I15) has formulas or values."""
        wb = get_workbook_formulas()
        ws = get_task_sheet(wb)
        csv_data = load_csv_data()

        has_data = 0
        for row in range(11, 16):
            for col in COLUMNS:
                cell = f"{col}{row}"
                if ws[cell].value is not None or csv_data.get(cell) is not None:
                    has_data += 1

        wb.close()
        assert has_data > 0, "No data found in GNI section (F11:I15)"

    def test_consumption_have_data(self):
        """Verify consumption section (F18:I22) has formulas or values."""
        wb = get_workbook_formulas()
        ws = get_task_sheet(wb)
        csv_data = load_csv_data()

        has_data = 0
        for row in range(18, 23):
            for col in COLUMNS:
                cell = f"{col}{row}"
                if ws[cell].value is not None or csv_data.get(cell) is not None:
                    has_data += 1

        wb.close()
        assert has_data > 0, "No data found in consumption section (F18:I22)"

    def test_population_have_data(self):
        """Verify population section (F25:I29) has formulas or values."""
        wb = get_workbook_formulas()
        ws = get_task_sheet(wb)
        csv_data = load_csv_data()

        has_data = 0
        for row in range(25, 30):
            for col in COLUMNS:
                cell = f"{col}{row}"
                if ws[cell].value is not None or csv_data.get(cell) is not None:
                    has_data += 1

        wb.close()
        assert has_data > 0, "No data found in population section (F25:I29)"


class TestStep1DataValues:
    """Test Step 1: Verify retrieved data values are correct."""

    def test_gni_values(self):
        """Verify GNI data values match expected."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        for row, expected in EXPECTED_GNI.items():
            for col_idx, col in enumerate(COLUMNS):
                cell = f"{col}{row}"
                actual = cell_value(ws, cell)
                exp = expected[col_idx]
                if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                    errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "GNI value mismatches:\n" + "\n".join(errors)

    def test_consumption_values(self):
        """Verify consumption data values match expected."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        for row, expected in EXPECTED_CONS.items():
            for col_idx, col in enumerate(COLUMNS):
                cell = f"{col}{row}"
                actual = cell_value(ws, cell)
                exp = expected[col_idx]
                if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                    errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "Consumption value mismatches:\n" + "\n".join(errors)

    def test_population_values(self):
        """Verify population data values match expected."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        for row, expected in EXPECTED_POP.items():
            for col_idx, col in enumerate(COLUMNS):
                cell = f"{col}{row}"
                actual = cell_value(ws, cell)
                exp = expected[col_idx]
                if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                    errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "Population value mismatches:\n" + "\n".join(errors)


class TestStep2SavingsRateCalculation:
    """Test Step 2: Savings rate calculations."""

    def test_savings_rate_have_data(self):
        """Verify savings rate cells (F33:I37) have formulas or values."""
        wb = get_workbook_formulas()
        ws = get_task_sheet(wb)
        csv_data = load_csv_data()

        has_data = 0
        for row in range(33, 38):
            for col in COLUMNS:
                cell = f"{col}{row}"
                if ws[cell].value is not None or csv_data.get(cell) is not None:
                    has_data += 1

        wb.close()
        assert has_data >= 16, f"Expected at least 16 cells with data in F33:I37, found {has_data}"

    def test_savings_rate_values(self):
        """Verify savings rate values are calculated correctly."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        for row, expected in EXPECTED_SAVINGS_RATE.items():
            for col_idx, col in enumerate(COLUMNS):
                cell = f"{col}{row}"
                actual = cell_value(ws, cell)
                exp = expected[col_idx]
                if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                    errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "Savings rate % mismatches:\n" + "\n".join(errors)


class TestStep2Statistics:
    """Test Step 2: Statistical calculations."""

    def test_statistics_have_data(self):
        """Verify statistics cells (F40:I46) have formulas or values."""
        wb = get_workbook_formulas()
        ws = get_task_sheet(wb)
        csv_data = load_csv_data()

        has_data = 0
        for row in range(40, 47):
            for col in COLUMNS:
                cell = f"{col}{row}"
                if ws[cell].value is not None or csv_data.get(cell) is not None:
                    has_data += 1

        wb.close()
        assert has_data >= 24, f"Expected at least 24 cells with data in F40:I46, found {has_data}"

    def test_min_values(self):
        """Verify MIN calculation values."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        expected = EXPECTED_STATS[40]
        for col_idx, col in enumerate(COLUMNS):
            cell = f"{col}40"
            actual = cell_value(ws, cell)
            exp = expected[col_idx]
            if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "MIN value mismatches:\n" + "\n".join(errors)

    def test_max_values(self):
        """Verify MAX calculation values."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        expected = EXPECTED_STATS[41]
        for col_idx, col in enumerate(COLUMNS):
            cell = f"{col}41"
            actual = cell_value(ws, cell)
            exp = expected[col_idx]
            if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "MAX value mismatches:\n" + "\n".join(errors)

    def test_median_values(self):
        """Verify MEDIAN calculation values."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        expected = EXPECTED_STATS[42]
        for col_idx, col in enumerate(COLUMNS):
            cell = f"{col}42"
            actual = cell_value(ws, cell)
            exp = expected[col_idx]
            if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "MEDIAN value mismatches:\n" + "\n".join(errors)

    def test_mean_values(self):
        """Verify AVERAGE calculation values."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        expected = EXPECTED_STATS[43]
        for col_idx, col in enumerate(COLUMNS):
            cell = f"{col}43"
            actual = cell_value(ws, cell)
            exp = expected[col_idx]
            if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "MEAN value mismatches:\n" + "\n".join(errors)

    def test_percentile_values(self):
        """Verify PERCENTILE calculation values (25th and 75th)."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        expected_25 = EXPECTED_STATS[44]
        for col_idx, col in enumerate(COLUMNS):
            cell = f"{col}44"
            actual = cell_value(ws, cell)
            exp = expected_25[col_idx]
            if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                errors.append(f"{cell} (25th): expected {exp}, got {actual}")

        expected_75 = EXPECTED_STATS[45]
        for col_idx, col in enumerate(COLUMNS):
            cell = f"{col}45"
            actual = cell_value(ws, cell)
            exp = expected_75[col_idx]
            if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                errors.append(f"{cell} (75th): expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "PERCENTILE value mismatches:\n" + "\n".join(errors)

    def test_stdev_values(self):
        """Verify STDEV calculation values."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        expected = EXPECTED_STATS[46]
        for col_idx, col in enumerate(COLUMNS):
            cell = f"{col}46"
            actual = cell_value(ws, cell)
            exp = expected[col_idx]
            if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "STDEV value mismatches:\n" + "\n".join(errors)


class TestStep3WeightedMean:
    """Test Step 3: Population-weighted mean calculation using SUMPRODUCT."""

    def test_weighted_mean_has_data(self):
        """Verify weighted mean row 49 has formulas or values."""
        wb = get_workbook_formulas()
        ws = get_task_sheet(wb)
        csv_data = load_csv_data()

        has_data = 0
        for col in COLUMNS:
            cell = f"{col}49"
            if ws[cell].value is not None or csv_data.get(cell) is not None:
                has_data += 1

        wb.close()
        assert has_data >= 3, f"Expected at least 3 cells with data in row 49, found {has_data}"

    def test_weighted_mean_values_reasonable(self):
        """Verify weighted mean values are numeric and in expected range."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        for _col_idx, col in enumerate(COLUMNS):
            cell = f"{col}49"
            actual = cell_value(ws, cell)

            if actual is None or actual == "" or not isinstance(actual, (int, float)):
                errors.append(f"{cell}: expected numeric value, got {actual}")
            elif actual == 0:
                errors.append(f"{cell}: value should not be zero")
            # Weighted mean should be in range 0-60 (weighted average of savings rate percentages)
            elif actual < 0 or actual > 60:
                errors.append(f"{cell}: value {actual} outside expected range (0-60)")

        wb.close()
        assert len(errors) == 0, "Weighted mean issues:\n" + "\n".join(errors)

    def test_weighted_mean_values(self):
        """Verify weighted mean values match expected."""
        wb = get_workbook()
        ws = get_task_sheet(wb)

        errors = []
        for col_idx, col in enumerate(COLUMNS):
            cell = f"{col}49"
            actual = cell_value(ws, cell)
            exp = EXPECTED_WEIGHTED_MEAN[col_idx]
            if actual is None or actual == "" or not isinstance(actual, (int, float)) or abs(actual - exp) > TOLERANCE:
                errors.append(f"{cell}: expected {exp}, got {actual}")

        wb.close()
        assert len(errors) == 0, "Weighted mean value mismatches:\n" + "\n".join(errors)


class TestNoExcelErrors:
    """Test that there are no Excel formula errors in the file."""

    def test_no_formula_errors(self):
        """Verify no #VALUE!, #REF!, #NAME?, etc. errors in calculated cells."""
        csv_data = load_csv_data()

        excel_errors = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
        error_cells = []

        # Check all relevant ranges
        ranges_to_check = [
            (11, 16),  # GNI
            (18, 23),  # Consumption
            (25, 30),  # Population
            (33, 38),  # Savings rate %
            (40, 47),  # Statistics
            (49, 50),  # Weighted mean
        ]

        for start_row, end_row in ranges_to_check:
            for row in range(start_row, end_row):
                for col in COLUMNS:
                    cell = f"{col}{row}"
                    val = csv_data.get(cell)
                    if val is not None and isinstance(val, str):
                        for err in excel_errors:
                            if err in str(val):
                                error_cells.append(f"{cell}: {val}")
                                break

        # Also check xlsx
        wb = get_workbook()
        ws = get_task_sheet(wb)
        for start_row, end_row in ranges_to_check:
            for row in range(start_row, end_row):
                for col in COLUMNS:
                    cell = f"{col}{row}"
                    val = ws[cell].value
                    if val is not None and isinstance(val, str):
                        for err in excel_errors:
                            if err in str(val):
                                if f"{cell}:" not in str(error_cells):
                                    error_cells.append(f"{cell}: {val}")
                                break
        wb.close()

        assert len(error_cells) == 0, "Excel errors found:\n" + "\n".join(error_cells)


EXPECTED_SHEETS = ["Task", "Data"]

EXPECTED_TASK_COLUMN_WIDTHS = {
    "A": 8.0,
    "B": 16.0,
    "C": 19.0,
    "D": 50.0,
    "E": 14.0,
    "F": 12.0,
    "G": 12.0,
    "H": 12.0,
    "I": 12.0,
}

EXPECTED_DATA_COLUMN_WIDTHS = {
    "A": 16.0,
    "B": 18.0,
    "C": 52.0,
    "D": 14.0,
    "E": 12.0,
    "F": 12.0,
    "G": 12.0,
    "H": 12.0,
}

WIDTH_TOLERANCE_RATIO = 0.5


class TestSheetStructure:
    """Test that no extra sheets were added to the workbook."""

    def test_no_extra_sheets_added(self):
        """Verify that only the expected sheets exist (Task and Data)."""
        wb = get_workbook()
        actual_sheets = wb.sheetnames
        wb.close()

        extra_sheets = [s for s in actual_sheets if s not in EXPECTED_SHEETS]

        assert len(extra_sheets) == 0, (
            f"Extra sheets were added to the workbook: {extra_sheets}. "
            f"Only {EXPECTED_SHEETS} should exist. "
            f"Found sheets: {actual_sheets}"
        )

    def test_required_sheets_exist(self):
        """Verify that the required sheets (Task and Data) still exist."""
        wb = get_workbook()
        actual_sheets = wb.sheetnames
        wb.close()

        missing_sheets = [s for s in EXPECTED_SHEETS if s not in actual_sheets]

        assert len(missing_sheets) == 0, (
            f"Required sheets are missing: {missing_sheets}. "
            f"Expected sheets: {EXPECTED_SHEETS}. "
            f"Found sheets: {actual_sheets}"
        )


class TestFormattingPreserved:
    """Test that formatting (column widths) was not drastically changed."""

    def test_task_sheet_column_widths(self):
        """Verify Task sheet column widths haven't been drastically changed."""
        wb = get_workbook()

        task_ws = None
        for sheet_name in wb.sheetnames:
            if "Task" in sheet_name or "task" in sheet_name.lower():
                task_ws = wb[sheet_name]
                break

        if task_ws is None:
            wb.close()
            pytest.skip("Task sheet not found")

        # ssconvert normalizes widths to ~1.7 as a conversion artifact—skip if detected
        widths = [task_ws.column_dimensions[col].width or 8.43 for col in EXPECTED_TASK_COLUMN_WIDTHS]
        if any(w < 3 for w in widths):
            wb.close()
            pytest.skip("Column widths appear to be normalized by ssconvert (conversion artifact)")

        errors = []
        for col, expected_width in EXPECTED_TASK_COLUMN_WIDTHS.items():
            actual_width = task_ws.column_dimensions[col].width
            if actual_width is None:
                actual_width = 8.43

            min_width = expected_width * (1 - WIDTH_TOLERANCE_RATIO)
            max_width = expected_width * (1 + WIDTH_TOLERANCE_RATIO)

            if actual_width < min_width or actual_width > max_width:
                errors.append(
                    f"Column {col}: expected width ~{expected_width:.1f}, "
                    f"got {actual_width:.1f} (outside {min_width:.1f}-{max_width:.1f} range)"
                )

        wb.close()
        assert len(errors) == 0, (
            "Task sheet column widths were drastically changed:\n"
            + "\n".join(errors)
            + "\nFormatting should be preserved for human readability."
        )

    def test_data_sheet_column_widths(self):
        """Verify Data sheet column widths haven't been drastically changed."""
        wb = get_workbook()

        data_ws = None
        for sheet_name in wb.sheetnames:
            if sheet_name == "Data":
                data_ws = wb[sheet_name]
                break

        if data_ws is None:
            wb.close()
            pytest.skip("Data sheet not found")

        # ssconvert normalizes widths to ~1.7 as a conversion artifact—skip if detected
        widths = [data_ws.column_dimensions[col].width or 8.43 for col in EXPECTED_DATA_COLUMN_WIDTHS]
        if any(w < 3 for w in widths):
            wb.close()
            pytest.skip("Column widths appear to be normalized by ssconvert (conversion artifact)")

        errors = []
        for col, expected_width in EXPECTED_DATA_COLUMN_WIDTHS.items():
            actual_width = data_ws.column_dimensions[col].width
            if actual_width is None:
                actual_width = 8.43

            min_width = expected_width * (1 - WIDTH_TOLERANCE_RATIO)
            max_width = expected_width * (1 + WIDTH_TOLERANCE_RATIO)

            if actual_width < min_width or actual_width > max_width:
                errors.append(
                    f"Column {col}: expected width ~{expected_width:.1f}, "
                    f"got {actual_width:.1f} (outside {min_width:.1f}-{max_width:.1f} range)"
                )

        wb.close()
        assert len(errors) == 0, (
            "Data sheet column widths were drastically changed:\n"
            + "\n".join(errors)
            + "\nFormatting should be preserved for human readability."
        )


class TestNoMacros:
    """Test that no VBA macros were introduced in the Excel file."""

    def test_no_vba_macros(self):
        """Verify the Excel file does not contain VBA macro code."""
        import zipfile

        with zipfile.ZipFile(EXCEL_FILE, "r") as zf:
            vba_files = [n for n in zf.namelist() if "vbaProject" in n or n.endswith(".bin")]

        assert len(vba_files) == 0, (
            f"The Excel file contains VBA macro code: {vba_files}. "
            "Macros are not allowed - please use only Excel formulas."
        )

    def test_file_extension_xlsx(self):
        """Verify the file is .xlsx (not .xlsm which supports macros)."""
        assert EXCEL_FILE.suffix.lower() == ".xlsx", (
            f"Expected .xlsx file extension, got {EXCEL_FILE.suffix}. "
            "Macro-enabled formats (.xlsm) are not allowed."
        )
