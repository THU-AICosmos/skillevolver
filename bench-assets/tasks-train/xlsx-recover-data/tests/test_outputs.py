"""
Tests for xlsx-recover-data training variant.

15 missing values across 4 sheets with dependency chains.
Tests use parameterization to group related checks.
"""
import pytest
import openpyxl
from pathlib import Path


@pytest.fixture
def workbook():
    """Load the recovered university enrollment workbook."""
    path = Path("university_enrollment_recovered.xlsx")
    if not path.exists():
        pytest.skip("university_enrollment_recovered.xlsx not found")
    return openpyxl.load_workbook(path)


@pytest.fixture
def enroll_sheet(workbook):
    return workbook["Enrollment by Department"]


@pytest.fixture
def growth_sheet(workbook):
    return workbook["Annual Growth (%)"]


@pytest.fixture
def mix_sheet(workbook):
    return workbook["Department Mix (%)"]


@pytest.fixture
def trend_sheet(workbook):
    return workbook["Trend Summary"]


def test_file_exists():
    """Test that the output file was created."""
    assert Path("university_enrollment_recovered.xlsx").exists()


# ==================== ALL 13 MISSING VALUES ====================

ENROLLMENT_VALUES = [
    # Tier 1
    ("J3", 7640, "AY2019 Total"),
    ("G7", 1400, "AY2023 Computing"),
    ("E5", 800, "AY2021 Arts"),
    # Tier 2
    ("E3", 790, "AY2019 Arts"),
    ("J5", 8180, "AY2021 Total"),
]

GROWTH_VALUES = [
    # Tier 2
    ("G6", 6.87, "AY2023 Computing Growth"),
    ("E5", 2.50, "AY2022 Arts Growth"),
    # Tier 3
    ("J4", 4.54, "AY2021 Total Growth"),
]

MIX_VALUES = [
    # Tier 2
    ("B3", 16.75, "AY2019 Engineering Mix"),
    # Tier 3
    ("E5", 9.78, "AY2021 Arts Mix"),
]

TREND_VALUES = [
    # Tier 1
    ("F4", 95, "Law 5yr Change"),
    # Tier 3
    ("E6", 1.71, "Arts CAGR"),
    ("E7", 790, "Arts Base Year"),
    # Bare "Avg Enrollment" row: sheet title declares "5-Year Trend Summary
    # (AY2019-AY2024)" so the Avg covers the historical 5 years AY2019-AY2023
    # and EXCLUDES the AY2024 endpoint (which is used for 5yr-Change/CAGR).
    # Same convention as neighbouring pre-filled B5/C5/D5/F5.
    ("E5", 804.0, "Arts Avg Enrollment (5-year, drop endpoint)"),
    # Labeled partial-window avg: mean of AY2020-AY2023 (4 of 5 years spanned
    # by the "4yr Change (AY2020-AY2024)" row, excluding AY2024 endpoint).
    ("D9", 1035.0, "Medicine Avg Enrollment (AY2020-AY2023)"),
]


# ==================== PARAMETERIZED TESTS (4 tests for 13 values) ====================

@pytest.mark.parametrize("cell,expected,desc", ENROLLMENT_VALUES)
def test_enrollment_values(enroll_sheet, cell, expected, desc):
    """Test all recovered Enrollment sheet values."""
    actual = enroll_sheet[cell].value
    assert actual == expected, f"{desc} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,desc", GROWTH_VALUES)
def test_growth_values(growth_sheet, cell, expected, desc):
    """Test all recovered Annual Growth sheet values."""
    actual = growth_sheet[cell].value
    assert abs(actual - expected) < 0.1, f"{desc} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,desc", MIX_VALUES)
def test_mix_values(mix_sheet, cell, expected, desc):
    """Test all recovered Department Mix sheet values."""
    actual = mix_sheet[cell].value
    assert abs(actual - expected) < 0.1, f"{desc} ({cell}): expected {expected}, got {actual}"


@pytest.mark.parametrize("cell,expected,desc", TREND_VALUES)
def test_trend_values(trend_sheet, cell, expected, desc):
    """Test all recovered Trend Summary sheet values."""
    actual = trend_sheet[cell].value
    if isinstance(expected, int):
        assert actual == expected, f"{desc} ({cell}): expected {expected}, got {actual}"
    else:
        assert abs(actual - expected) < 0.5, f"{desc} ({cell}): expected {expected}, got {actual}"


# ==================== VALIDATION TESTS ====================

def test_no_remaining_placeholders(workbook):
    """Test that no '???' placeholders remain in any sheet."""
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell == "???":
                    pytest.fail(f"Found '???' in '{sheet_name}'")


def test_row_sums_consistent(enroll_sheet):
    """Verify recovered totals match row sums."""
    # AY2019 (row 3)
    row3_sum = sum(enroll_sheet.cell(row=3, column=col).value for col in range(2, 10))
    assert row3_sum == enroll_sheet['J3'].value, "AY2019 sum mismatch"

    # AY2021 (row 5)
    row5_sum = sum(enroll_sheet.cell(row=5, column=col).value for col in range(2, 10))
    assert row5_sum == enroll_sheet['J5'].value, "AY2021 sum mismatch"


def test_cross_sheet_consistency(enroll_sheet, trend_sheet):
    """Verify cross-sheet values are consistent."""
    # CAGR should match calculation from Enrollment
    start = enroll_sheet['E3'].value  # AY2019 Arts
    end = enroll_sheet['E8'].value    # AY2024 Arts
    expected_cagr = round(((end / start) ** 0.2 - 1) * 100, 2)
    actual_cagr = trend_sheet['E6'].value
    assert abs(actual_cagr - expected_cagr) < 0.1, f"CAGR mismatch: {expected_cagr} vs {actual_cagr}"

    # Base year should match enrollment
    assert trend_sheet['E7'].value == enroll_sheet['E3'].value, "E7 should equal E3"
